#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel比对工具 - UI界面版
Python 3.7.1 兼容
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from decimal import Decimal, ROUND_HALF_UP

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelCompareApp:
    """Excel比对工具主应用"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Excel比对工具 v1.0")
        self.root.geometry("680x600")
        self.root.minsize(600, 550)
        
        # 变量
        self.work_dir = tk.StringVar(value=os.getcwd())
        self.base_file = tk.StringVar()
        self.data_a_file = tk.StringVar()
        self.data_b_file = tk.StringVar()
        self.output_file = tk.StringVar(value="compare_result.xlsx")
        
        # 阈值 - 默认值
        self.green_threshold = tk.StringVar(value="1.0")
        self.red_min_threshold = tk.StringVar(value="1.0")
        self.red_max_threshold = tk.StringVar(value="100.0")
        
        self.create_ui()
        
    def create_ui(self):
        """创建UI - 使用grid布局"""
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # 主容器
        main = tk.Frame(self.root, padx=20, pady=15)
        main.grid(row=0, column=0, sticky="nsew")
        main.columnconfigure(0, weight=1)
        
        current_row = 0
        
        # ===== 标题 =====
        title = tk.Label(main, text="Excel 数据比对工具", font=("Helvetica", 18, "bold"))
        title.grid(row=current_row, column=0, pady=(0, 20), sticky="w")
        current_row += 1
        
        # ===== 1. 工作目录 =====
        sec1 = tk.LabelFrame(main, text=" 工作目录 ", font=("Helvetica", 11), padx=15, pady=12)
        sec1.grid(row=current_row, column=0, sticky="ew", pady=(0, 12))
        sec1.columnconfigure(1, weight=1)
        current_row += 1
        
        tk.Label(sec1, text="目录路径:", font=("Helvetica", 11)).grid(row=0, column=0, sticky="w", padx=(0, 10))
        tk.Entry(sec1, textvariable=self.work_dir, font=("Helvetica", 11)).grid(row=0, column=1, sticky="ew", padx=(0, 10))
        tk.Button(sec1, text="选择目录", command=self.browse_workdir, width=10).grid(row=0, column=2)
        
        # ===== 2. 文件选择 =====
        sec2 = tk.LabelFrame(main, text=" 文件选择 ", font=("Helvetica", 11), padx=15, pady=12)
        sec2.grid(row=current_row, column=0, sticky="ew", pady=(0, 12))
        sec2.columnconfigure(1, weight=1)
        current_row += 1
        
        # 基准文件
        tk.Label(sec2, text="上传基准文件:", font=("Helvetica", 11)).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=4)
        tk.Entry(sec2, textvariable=self.base_file, font=("Helvetica", 11)).grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=4)
        tk.Button(sec2, text="浏览...", command=lambda: self.browse_file(self.base_file), width=10).grid(row=0, column=2, pady=4)
        
        # 输入1
        tk.Label(sec2, text="上传输入1文件:", font=("Helvetica", 11)).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=4)
        tk.Entry(sec2, textvariable=self.data_a_file, font=("Helvetica", 11)).grid(row=1, column=1, sticky="ew", padx=(0, 10), pady=4)
        tk.Button(sec2, text="浏览...", command=lambda: self.browse_file(self.data_a_file), width=10).grid(row=1, column=2, pady=4)
        
        # 输入2
        tk.Label(sec2, text="上传输入2文件:", font=("Helvetica", 11)).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=4)
        tk.Entry(sec2, textvariable=self.data_b_file, font=("Helvetica", 11)).grid(row=2, column=1, sticky="ew", padx=(0, 10), pady=4)
        tk.Button(sec2, text="浏览...", command=lambda: self.browse_file(self.data_b_file), width=10).grid(row=2, column=2, pady=4)
        
        # 输出文件名
        tk.Label(sec2, text="输出文件名:", font=("Helvetica", 11)).grid(row=3, column=0, sticky="w", padx=(0, 10), pady=4)
        tk.Entry(sec2, textvariable=self.output_file, font=("Helvetica", 11)).grid(row=3, column=1, sticky="ew", padx=(0, 10), pady=4)
        
        # ===== 3. 颜色阈值 =====
        sec3 = tk.LabelFrame(main, text=" 颜色阈值设置 (差异百分比绝对值) ", font=("Helvetica", 11), padx=15, pady=12)
        sec3.grid(row=current_row, column=0, sticky="ew", pady=(0, 12))
        current_row += 1
        
        # 绿色行
        row_green = tk.Frame(sec3)
        row_green.grid(row=0, column=0, sticky="w", pady=4)
        tk.Label(row_green, text="   ", bg="#90EE90", width=3, relief="solid", bd=1).grid(row=0, column=0)
        tk.Label(row_green, text="  绿色:  差异 < ", font=("Helvetica", 11)).grid(row=0, column=1)
        tk.Entry(row_green, textvariable=self.green_threshold, width=8, font=("Helvetica", 11)).grid(row=0, column=2)
        tk.Label(row_green, text=" %", font=("Helvetica", 11)).grid(row=0, column=3)
        
        # 红色行
        row_red = tk.Frame(sec3)
        row_red.grid(row=1, column=0, sticky="w", pady=4)
        tk.Label(row_red, text="   ", bg="#FF6B6B", width=3, relief="solid", bd=1).grid(row=0, column=0)
        tk.Label(row_red, text="  红色:  ", font=("Helvetica", 11)).grid(row=0, column=1)
        tk.Entry(row_red, textvariable=self.red_min_threshold, width=8, font=("Helvetica", 11)).grid(row=0, column=2)
        tk.Label(row_red, text=" % <= 差异 <= ", font=("Helvetica", 11)).grid(row=0, column=3)
        tk.Entry(row_red, textvariable=self.red_max_threshold, width=8, font=("Helvetica", 11)).grid(row=0, column=4)
        tk.Label(row_red, text=" %", font=("Helvetica", 11)).grid(row=0, column=5)
        
        # 无色行
        row_white = tk.Frame(sec3)
        row_white.grid(row=2, column=0, sticky="w", pady=4)
        tk.Label(row_white, text="   ", bg="#FFFFFF", width=3, relief="solid", bd=1).grid(row=0, column=0)
        tk.Label(row_white, text="  无色:  差异 > 红色上限", font=("Helvetica", 11)).grid(row=0, column=1)
        
        # ===== 4. 操作按钮 =====
        sec4 = tk.Frame(main)
        sec4.grid(row=current_row, column=0, sticky="ew", pady=(0, 12))
        current_row += 1
        
        tk.Button(sec4, text="生成测试文件", command=self.generate_test_files, 
                  width=14, height=2, font=("Helvetica", 10)).grid(row=0, column=0, padx=(0, 10))
        
        tk.Button(sec4, text="开始对比", command=self.run_compare,
                  width=14, height=2, bg="#4CAF50", fg="white", 
                  font=("Helvetica", 11, "bold")).grid(row=0, column=1, padx=(0, 10))
        
        tk.Button(sec4, text="打开结果", command=self.open_result, 
                  width=12, height=2, font=("Helvetica", 10)).grid(row=0, column=2, padx=(0, 10))
        
        tk.Button(sec4, text="打开目录", command=self.open_workdir,
                  width=12, height=2, font=("Helvetica", 10)).grid(row=0, column=3)
        
        # ===== 5. 日志 =====
        sec5 = tk.LabelFrame(main, text=" 运行日志 ", font=("Helvetica", 11), padx=10, pady=8)
        sec5.grid(row=current_row, column=0, sticky="nsew", pady=(0, 5))
        sec5.columnconfigure(0, weight=1)
        sec5.rowconfigure(0, weight=1)
        main.rowconfigure(current_row, weight=1)
        
        self.log_text = tk.Text(sec5, height=6, font=("Courier", 10), wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        
        scrollbar = tk.Scrollbar(sec5, command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # 初始日志
        self.log("欢迎使用Excel比对工具!")
        self.log("步骤: 1.选择目录 -> 2.上传文件 -> 3.点击开始对比")
        if not OPENPYXL_AVAILABLE:
            self.log("[错误] 缺少openpyxl库: pip install openpyxl")
    
    # ========== 辅助方法 ==========
    
    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def browse_workdir(self):
        d = filedialog.askdirectory(title="选择工作目录", initialdir=self.work_dir.get())
        if d:
            self.work_dir.set(d)
            self.log("工作目录: " + d)
            
    def browse_file(self, var):
        f = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=self.work_dir.get(),
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")]
        )
        if f:
            var.set(f)
            
    def open_result(self):
        path = os.path.join(self.work_dir.get(), self.output_file.get())
        if os.path.exists(path):
            if sys.platform == 'darwin':
                os.system('open "{}"'.format(path))
            elif sys.platform == 'win32':
                os.system('start "" "{}"'.format(path))
            else:
                os.system('xdg-open "{}"'.format(path))
        else:
            messagebox.showwarning("提示", "结果文件不存在")
            
    def open_workdir(self):
        d = self.work_dir.get()
        if os.path.exists(d):
            if sys.platform == 'darwin':
                os.system('open "{}"'.format(d))
            elif sys.platform == 'win32':
                os.system('explorer "{}"'.format(d))
            else:
                os.system('xdg-open "{}"'.format(d))
    
    # ========== 核心功能 ==========
    
    def generate_test_files(self):
        """生成测试文件"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("错误", "缺少openpyxl库")
            return
            
        workdir = self.work_dir.get()
        if not os.path.exists(workdir):
            messagebox.showerror("错误", "工作目录不存在")
            return
            
        self.log("=" * 40)
        self.log("生成测试文件...")
        
        try:
            # 基准文件
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="指标名称")
            indicators = [
                "正常_完全相同", "正常_小差异_0.5%", "正常_临界_1%", "正常_中等_5%",
                "正常_较大_50%", "正常_超大_150%", "特殊_B为零", "特殊_负数",
                "缺失_A无数据", "缺失_B无数据", "缺失_都无数据"
            ]
            for i, name in enumerate(indicators, 2):
                ws.cell(row=i, column=1, value=name)
            wb.save(os.path.join(workdir, "test_base.xlsx"))
            self.log("  test_base.xlsx")
            
            # 数据A
            wb = Workbook()
            ws = wb.active
            data_a = [
                ("正常_完全相同", 1000000), ("正常_小差异_0.5%", 1005000),
                ("正常_临界_1%", 1010000), ("正常_中等_5%", 1050000),
                ("正常_较大_50%", 1500000), ("正常_超大_150%", 2500000),
                ("特殊_B为零", 100), ("特殊_负数", -500),
                ("缺失_A无数据", None), ("缺失_都无数据", None)
            ]
            for col, (h, v) in enumerate(data_a, 1):
                ws.cell(row=1, column=col, value=h)
                ws.cell(row=2, column=col, value=v)
            wb.save(os.path.join(workdir, "test_data_a.xlsx"))
            self.log("  test_data_a.xlsx")
            
            # 数据B
            wb = Workbook()
            ws = wb.active
            data_b = [
                ("正常_完全相同", 1000000), ("正常_小差异_0.5%", 1000000),
                ("正常_临界_1%", 1000000), ("正常_中等_5%", 1000000),
                ("正常_较大_50%", 1000000), ("正常_超大_150%", 1000000),
                ("特殊_B为零", 0), ("特殊_负数", -400),
                ("缺失_B无数据", None), ("缺失_都无数据", None)
            ]
            for col, (h, v) in enumerate(data_b, 1):
                ws.cell(row=1, column=col, value=h)
                ws.cell(row=2, column=col, value=v)
            wb.save(os.path.join(workdir, "test_data_b.xlsx"))
            self.log("  test_data_b.xlsx")
            
            # 自动填充路径
            self.base_file.set(os.path.join(workdir, "test_base.xlsx"))
            self.data_a_file.set(os.path.join(workdir, "test_data_a.xlsx"))
            self.data_b_file.set(os.path.join(workdir, "test_data_b.xlsx"))
            
            self.log("[完成] 测试文件已生成并自动填充路径")
            messagebox.showinfo("成功", "测试文件已生成!")
            
        except Exception as e:
            self.log("[错误] " + str(e))
            messagebox.showerror("错误", str(e))
    
    def run_compare(self):
        """运行对比"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("错误", "缺少openpyxl库")
            return
            
        # 验证
        if not self.base_file.get():
            messagebox.showwarning("提示", "请上传基准文件")
            return
        if not self.data_a_file.get():
            messagebox.showwarning("提示", "请上传输入1文件")
            return
        if not self.data_b_file.get():
            messagebox.showwarning("提示", "请上传输入2文件")
            return
            
        try:
            green_th = float(self.green_threshold.get())
            red_min = float(self.red_min_threshold.get())
            red_max = float(self.red_max_threshold.get())
        except ValueError:
            messagebox.showwarning("提示", "阈值必须是数字")
            return
            
        self.log("=" * 40)
        self.log("开始对比...")
        self.log("阈值: 绿<{}%, 红{}%-{}%".format(green_th, red_min, red_max))
        
        try:
            # 读取基准
            base_names = self._read_base(self.base_file.get())
            self.log("基准: {} 个指标".format(len(base_names)))
            
            # 读取数据
            data_a = self._read_horizontal(self.data_a_file.get())
            self.log("输入1: {} 个数据".format(len(data_a)))
            
            data_b = self._read_horizontal(self.data_b_file.get())
            self.log("输入2: {} 个数据".format(len(data_b)))
            
            # 生成结果
            output = os.path.join(self.work_dir.get(), self.output_file.get())
            self._create_result(output, base_names, data_a, data_b, green_th, red_min, red_max)
            
            self.log("[完成] 结果已保存: " + self.output_file.get())
            messagebox.showinfo("成功", "对比完成!")
            
        except Exception as e:
            self.log("[错误] " + str(e))
            messagebox.showerror("错误", str(e))
            import traceback
            traceback.print_exc()
    
    def _read_base(self, path):
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        names = []
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=1).value
            if v:
                names.append(str(v).strip())
        wb.close()
        return names
        
    def _read_horizontal(self, path):
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        data = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                data[str(h).strip()] = ws.cell(row=2, column=col).value
        wb.close()
        return data
        
    def _parse_num(self, v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        s = str(v).strip().replace(',', '').replace(' ', '')
        if not s or s.lower() in ['error', '#value!', 'none', 'null']:
            return None
        try:
            return Decimal(s)
        except:
            return None
            
    def _create_result(self, output, names, data_a, data_b, green_th, red_min, red_max):
        GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        HEADER = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "比对结果"
        
        # 表头
        for col, h in enumerate(["指标名称", "A", "B", "差额", "差异%"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HEADER
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = border
            
        # 数据
        for row, name in enumerate(names, 2):
            ws.cell(row=row, column=1, value=name).border = border
            
            va = data_a.get(name)
            vb = data_b.get(name)
            pa = self._parse_num(va)
            pb = self._parse_num(vb)
            
            # A列
            if pa is not None:
                ws.cell(row=row, column=2, value=float(pa)).border = border
            else:
                ws.cell(row=row, column=2, value="error").border = border
                
            # B列
            if pb is not None:
                ws.cell(row=row, column=3, value=float(pb)).border = border
            else:
                ws.cell(row=row, column=3, value="error").border = border
                
            # 差额
            if pa is not None and pb is not None:
                diff = pa - pb
                ws.cell(row=row, column=4, value=float(diff.quantize(Decimal('0.0001')))).border = border
            else:
                ws.cell(row=row, column=4, value="#VALUE!").border = border
                diff = None
                
            # 差异%
            cell = ws.cell(row=row, column=5)
            cell.border = border
            if diff is not None and pb is not None and pb != 0:
                pct = (diff / pb) * 100
                pct_val = float(pct.quantize(Decimal('0.0001')))
                cell.value = "{}%".format(pct_val)
                
                abs_pct = abs(pct)
                if abs_pct < green_th:
                    cell.fill = GREEN
                elif red_min <= abs_pct <= red_max:
                    cell.fill = RED
            else:
                cell.value = "#VALUE!"
                
        # 列宽
        for col, w in enumerate([22, 18, 18, 16, 16], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
            
        # 图例
        lr = len(names) + 3
        ws.cell(row=lr, column=4, value="差异 >= {}%".format(red_min))
        ws.cell(row=lr, column=5, value="红色").fill = RED
        ws.cell(row=lr+1, column=4, value="差异 < {}%".format(green_th))
        ws.cell(row=lr+1, column=5, value="绿色").fill = GREEN
        
        wb.save(output)


def main():
    os.environ['TK_SILENCE_DEPRECATION'] = '1'
    root = tk.Tk()
    ExcelCompareApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
