#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试标准化功能：解析表头、维度聚合、指标匹配时都应该忽略下划线、空格和括号
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_test_files():
    """创建测试文件：A表和B表使用不同的列名格式（带下划线、空格、括号等）"""
    
    base_dir = '/Users/li.wang/ai-test-project/excel-tool'
    
    # 创建测试表A
    wb_a = Workbook()
    ws_a = wb_a.active
    ws_a.title = 'Sheet1'
    
    # 表A的列名：维度1, 维度_2, 指标 1, 指标(2), 指标_3
    headers_a = ['维度1', '维度_2', '指标 1', '指标(2)', '指标_3']
    ws_a.append(headers_a)
    
    # 表A的数据
    data_a = [
        ['地区A', '类型_X', 100, 200, 300],
        ['地区_B', '类型 Y', 150, 250, 350],
        ['地区(C)', '类型（Z）', 200, 300, 400],
    ]
    for row in data_a:
        ws_a.append(row)
    
    file_a = os.path.join(base_dir, 'test_table_a_normalize.xlsx')
    wb_a.save(file_a)
    print(f'创建测试表A: {file_a}')
    print(f'  列名: {headers_a}')
    
    # 创建测试表B - 使用不同的列名格式，但标准化后应该匹配
    wb_b = Workbook()
    ws_b = wb_b.active
    ws_b.title = 'Sheet1'
    
    # 表B的列名：维度 1（带空格）, 维度2（无下划线）, 指标1（无空格）, 指标_2（下划线）, 指标【4】（新指标）
    headers_b = ['维度 1', '维度2', '指标1', '指标_2', '指标【4】']
    ws_b.append(headers_b)
    
    # 表B的数据 - 维度值也使用不同的格式
    data_b = [
        ['地区 A', '类型X', 110, 210, 500],  # 地区A vs 地区 A, 类型_X vs 类型X
        ['地区B', '类型_Y', 160, 260, 600],   # 地区_B vs 地区B, 类型 Y vs 类型_Y
        ['地区【C】', '类型Z', 210, 310, 700],  # 地区(C) vs 地区【C】, 类型（Z）vs 类型Z
    ]
    for row in data_b:
        ws_b.append(row)
    
    file_b = os.path.join(base_dir, 'test_table_b_normalize.xlsx')
    wb_b.save(file_b)
    print(f'创建测试表B: {file_b}')
    print(f'  列名: {headers_b}')
    
    print('\n预期效果：')
    print('1. 表头解析时：')
    print('   - "维度1" 和 "维度 1" 应该被识别为同一列')
    print('   - "维度_2" 和 "维度2" 应该被识别为同一列')
    print('   - "指标 1" 和 "指标1" 应该被识别为同一列')
    print('   - "指标(2)" 和 "指标_2" 应该被识别为同一列')
    print('   - "指标_3" 是A表独有')
    print('   - "指标【4】" 是B表独有')
    print('   - 解析结果应该是去重后的5个列名')
    
    print('\n2. 聚合时：')
    print('   - 维度键匹配：["地区A", "类型_X"] 应该匹配 ["地区 A", "类型X"]')
    print('   - 维度键匹配：["地区_B", "类型 Y"] 应该匹配 ["地区B", "类型_Y"]')
    print('   - 维度键匹配：["地区(C)", "类型（Z）"] 应该匹配 ["地区【C】", "类型Z"]')
    print('   - 因此，聚合后应该有3行（对应3个维度组合）')
    
    print('\n3. 公共指标：')
    print('   - "指标 1/指标1" 和 "指标(2)/指标_2" 应该被识别为公共指标')
    print('   - Sheet3应该只包含这两个公共指标')
    
    return file_a, file_b

if __name__ == '__main__':
    create_test_files()

