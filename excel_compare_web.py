#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel比对工具 - Web界面版
Python 3.7.1 兼容
启动后在浏览器中打开 http://localhost:8080
"""

import os
import sys
import json
import webbrowser
import threading
import subprocess
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, unquote
from decimal import Decimal, ROUND_HALF_UP

# 设置标准输出编码为UTF-8（解决Windows控制台中文输出问题）
if sys.platform == 'win32':
    try:
        # Python 3.7+
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        # Python 3.6及更早版本
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# 设置环境变量
os.environ['PYTHONIOENCODING'] = 'utf-8'

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import xlrd
    XLRD_OK = True
except ImportError:
    XLRD_OK = False

# 全局配置
WORK_DIR = os.getcwd()
PORT = 9527

HTML_TEMPLATE = '''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Excel比对工具</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh; padding: 30px;
        }
        .container { 
            max-width: 800px; margin: 0 auto; 
            background: white; border-radius: 16px; 
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 35px; 
        }
        h1 { 
            text-align: center; color: #333; margin-bottom: 30px;
            font-size: 28px; font-weight: 600;
        }
        
        /* Tab 样式 */
        .tabs {
            display: flex;
            border-bottom: 2px solid #e0e0e0;
            margin-bottom: 25px;
            gap: 10px;
        }
        .tab {
            padding: 12px 24px;
            cursor: pointer;
            border-bottom: 3px solid transparent;
            transition: all 0.3s;
            font-size: 15px;
            font-weight: 500;
            color: #666;
        }
        .tab:hover { color: #667eea; }
        .tab.active { 
            border-bottom-color: #667eea; 
            color: #667eea;
            font-weight: 600; 
        }
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        
        .section { 
            background: #f8f9fa; border-radius: 12px; 
            padding: 20px; margin-bottom: 20px; 
        }
        .section-title { 
            font-size: 15px; font-weight: 600; color: #555; 
            margin-bottom: 15px; display: flex; align-items: center;
        }
        .section-title::before {
            content: ""; width: 4px; height: 18px; 
            background: #667eea; border-radius: 2px; margin-right: 10px;
        }
        .form-row { 
            display: flex; align-items: center; margin-bottom: 12px; 
        }
        .form-row:last-child { margin-bottom: 0; }
        label { 
            width: 130px; font-size: 14px; color: #444; font-weight: 500;
        }
        input[type="text"], input[type="file"] { 
            flex: 1; padding: 10px 14px; border: 2px solid #e0e0e0; 
            border-radius: 8px; font-size: 14px; transition: border-color 0.2s;
        }
        input[type="text"]:focus { border-color: #667eea; outline: none; }
        input[type="number"] {
            width: 80px; padding: 8px 12px; border: 2px solid #e0e0e0;
            border-radius: 8px; font-size: 14px; text-align: center;
        }
        .color-row { display: flex; align-items: center; margin-bottom: 10px; }
        .color-box { 
            width: 24px; height: 24px; border-radius: 4px; 
            margin-right: 12px; border: 1px solid #ccc;
        }
        .green-box { background: #90EE90; }
        .red-box { background: #FF6B6B; }
        .white-box { background: #fff; }
        .color-text { font-size: 14px; color: #555; }
        .btn-row { 
            display: flex; gap: 12px; margin-top: 25px; flex-wrap: wrap;
        }
        button { 
            padding: 12px 24px; border: none; border-radius: 8px; 
            font-size: 14px; font-weight: 600; cursor: pointer; 
            transition: transform 0.1s, box-shadow 0.2s;
        }
        button:hover { transform: translateY(-1px); }
        button:active { transform: translateY(0); }
        .btn-primary { 
            background: linear-gradient(135deg, #667eea, #764ba2); 
            color: white; box-shadow: 0 4px 15px rgba(102,126,234,0.4);
        }
        .btn-secondary { background: #6c757d; color: white; }
        .btn-success { background: #28a745; color: white; }
        .log-box { 
            background: #1e1e1e; color: #0f0; border-radius: 8px; 
            padding: 15px; font-family: "Courier New", monospace; 
            font-size: 13px; height: 150px; overflow-y: auto;
            white-space: pre-wrap;
        }
        .file-input-wrapper {
            flex: 1; display: flex; gap: 8px;
        }
        .file-path {
            flex: 1; padding: 10px 14px; border: 2px solid #e0e0e0;
            border-radius: 8px; font-size: 13px; background: #fff;
            overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
        }
        .btn-browse {
            padding: 10px 16px; background: #e9ecef; border: 2px solid #e0e0e0;
            border-radius: 8px; cursor: pointer; font-size: 13px;
        }
        .btn-browse:hover { background: #dee2e6; }
        .hidden-input { display: none; }
        .hint-text {
            font-size: 12px;
            color: #888;
            margin-top: 8px;
            font-style: italic;
        }
        
        /* Loading 样式 */
        .loading-overlay {
            display: none;
            position: fixed;
            top: 0; left: 0; right: 0; bottom: 0;
            background: rgba(0, 0, 0, 0.7);
            z-index: 9999;
            justify-content: center;
            align-items: center;
        }
        .loading-overlay.show { display: flex; }
        .loading-content {
            background: white;
            padding: 40px;
            border-radius: 16px;
            text-align: center;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
        }
        .spinner {
            width: 50px;
            height: 50px;
            margin: 0 auto 20px;
            border: 4px solid #f3f3f3;
            border-top: 4px solid #667eea;
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .loading-text {
            font-size: 16px;
            color: #333;
            font-weight: 500;
        }
        
        /* 多选框样式 */
        .multiselect-container {
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            padding: 10px;
            max-height: 200px;
            overflow-y: auto;
            background: white;
        }
        .multiselect-option {
            padding: 6px 10px;
            cursor: pointer;
            border-radius: 4px;
            margin-bottom: 4px;
            display: flex;
            align-items: center;
            font-size: 14px;
        }
        .multiselect-option:hover {
            background: #f0f0f0;
        }
        .multiselect-option input[type="checkbox"] {
            margin-right: 8px;
        }
        .column-selection-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
            margin-top: 15px;
        }
        .column-group {
            flex: 1;
        }
        .column-group-title {
            font-size: 14px;
            font-weight: 600;
            color: #555;
            margin-bottom: 8px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Excel 数据比对工具</h1>
        
        <!-- Tab 切换栏 -->
        <div class="tabs">
            <div class="tab active" onclick="switchTab(1)">指标比对</div>
            <div class="tab" onclick="switchTab(2)">指标+维度比对</div>
            <div class="tab" onclick="switchTab(3)">聚合+比对</div>
        </div>
        
        <!-- Tab 1: 指标比对 -->
        <div id="tab1-content" class="tab-content active">
            <div class="section">
                <div class="section-title">工作目录</div>
                <div class="form-row">
                    <label>目录路径:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="workDir" value="''' + WORK_DIR.replace('\\', '\\\\').replace("'", "\\'") + '''">
                        <button class="btn-browse" onclick="browseDir()">选择目录</button>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">文件选择</div>
                <div class="form-row">
                    <label>上传基准文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="baseFile" placeholder="选择基准匹配列文件 (.xlsx)">
                        <button class="btn-browse" onclick="browseFile('baseFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>上传输入1文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="dataAFile" placeholder="选择输入1数据文件 (.xlsx)">
                        <button class="btn-browse" onclick="browseFile('dataAFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>上传输入2文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="dataBFile" placeholder="选择输入2数据文件 (.xlsx)">
                        <button class="btn-browse" onclick="browseFile('dataBFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>输出文件名:</label>
                    <input type="text" id="outputFile" value="compare_result.xlsx">
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">比对设置</div>
                <div class="form-row">
                    <label>小数位数:</label>
                    <input type="number" id="decimalPlaces" value="6" min="0" max="10" step="1" 
                           style="width: 80px; margin: 0 10px;">
                    <span class="color-text">位（用于差额和百分比）</span>
                </div>
                <div class="form-row" style="margin-top: 10px;">
                    <label>阈值设置:</label>
                    <span class="color-text">百分比绝对值 < </span>
                    <input type="number" id="greenTh" value="1.0" step="0.1" style="width: 80px; margin: 0 5px;">
                    <span class="color-text">% 或 A=B 时为绿色，否则为红色</span>
                </div>
                <div class="color-row" style="margin-top: 10px;">
                    <div class="color-box green-box"></div>
                    <span class="color-text" style="margin-left: 10px;">绿色: A=B 或 |差异%| < 阈值</span>
                </div>
                <div class="color-row">
                    <div class="color-box red-box"></div>
                    <span class="color-text" style="margin-left: 10px;">红色: 其他情况</span>
                </div>
            </div>
            
            <div class="btn-row">
                <button class="btn-secondary" onclick="generateTest()">生成测试文件</button>
                <button class="btn-primary" onclick="runCompare()">🚀 开始对比</button>
                <button class="btn-success" onclick="openResult()">打开结果</button>
                <button class="btn-secondary" onclick="openDir()">打开目录</button>
            </div>
        </div>
        
        <!-- Tab 2: 指标+维度比对 -->
        <div id="tab2-content" class="tab-content">
            <div class="section">
                <div class="section-title">工作目录</div>
                <div class="form-row">
                    <label>目录路径:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="workDir2" value="''' + WORK_DIR.replace('\\', '\\\\').replace("'", "\\'") + '''">
                        <button class="btn-browse" onclick="browseDir2()">选择目录</button>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">文件选择</div>
                <div class="form-row">
                    <label>上传表A文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="tableAFile" placeholder="选择表A数据文件 (.xlsx)">
                        <button class="btn-browse" onclick="browseFile2('tableAFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>上传基准表（表B）:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="tableBFile" placeholder="选择基准表数据文件 (.xlsx)">
                        <button class="btn-browse" onclick="browseFile2('tableBFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>输出文件名:</label>
                    <input type="text" id="outputFile2" value="dimension_compare_result.xlsx">
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">比对设置</div>
                <div class="form-row">
                    <label>基准列数量:</label>
                    <input type="number" id="keyColumns" value="1" min="1" max="10" step="1" 
                           style="width: 80px; margin: 0 10px;">
                    <span class="color-text">列（前N列作为维度列进行匹配）</span>
                </div>
                <div class="form-row">
                    <label>差异阈值:</label>
                    <input type="number" id="diffThreshold" value="1" min="0" step="0.1" 
                           style="width: 80px; margin: 0 10px;">
                    <span class="color-text">（差异值绝对值 &lt; 阈值为绿色，≥ 阈值为红色）</span>
                </div>
                <div class="hint-text">
                    说明：<br>
                    1. 以前N列为基准进行行匹配（忽略空格、下划线、中英文括号差异）<br>
                    2. 指标列以B表为基准，只保留B表有的指标列<br>
                    3. 每个指标列显示差异值（A - B），根据阈值标记颜色<br>
                    4. 不匹配的行标记为"{文件名}表error"
                </div>
            </div>
            
            <div class="btn-row">
                <button class="btn-secondary" onclick="generateDimensionTest()">生成测试文件</button>
                <button class="btn-primary" onclick="runDimensionCompare()">🚀 开始对比</button>
                <button class="btn-success" onclick="openDimensionResult()">打开结果</button>
                <button class="btn-secondary" onclick="openDir2()">打开目录</button>
            </div>
        </div>
        
        <!-- Tab 3: 聚合+比对 -->
        <div id="tab3-content" class="tab-content">
            <div class="section">
                <div class="section-title">工作目录</div>
                <div class="form-row">
                    <label>目录路径:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="workDir3" value="''' + WORK_DIR.replace('\\', '\\\\').replace("'", "\\'") + '''">
                        <button class="btn-browse" onclick="browseDir3()">选择目录</button>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">文件上传</div>
                <div class="form-row">
                    <label>表A文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="aggTableAFile" placeholder="选择表A文件 (.xlsx, .xls)">
                        <button class="btn-browse" onclick="browseFile('aggTableAFile')">选择文件</button>
                    </div>
                </div>
                <div class="form-row">
                    <label>表B文件:</label>
                    <div class="file-input-wrapper">
                        <input type="text" id="aggTableBFile" placeholder="选择表B文件 (.xlsx, .xls)">
                        <button class="btn-browse" onclick="browseFile('aggTableBFile')">选择文件</button>
                    </div>
                </div>
                <div class="btn-row">
                    <button class="btn-primary" onclick="parseHeaders()">🔍 解析维度与指标名</button>
                </div>
            </div>
            
            <div class="section" id="columnSelectionSection" style="display: none;">
                <div class="section-title">列选择（从合并的表头中选择）</div>
                <div class="column-selection-row">
                    <div class="column-group">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                            <div class="column-group-title" style="margin-bottom: 0;">维度列（用于分组）</div>
                            <button class="btn-browse" style="padding: 4px 12px; font-size: 12px;" onclick="toggleAllDimensions()">全选/取消</button>
                        </div>
                        <div class="multiselect-container" id="dimensionColumns"></div>
                    </div>
                    <div class="column-group">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                            <div class="column-group-title" style="margin-bottom: 0;">指标列（用于聚合求和）</div>
                            <button class="btn-browse" style="padding: 4px 12px; font-size: 12px;" onclick="toggleAllIndicators()">全选/取消</button>
                        </div>
                        <div class="multiselect-container" id="indicatorColumns"></div>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">比对参数</div>
                <div class="form-row">
                    <label>差异阈值:</label>
                    <input type="number" id="aggDiffThreshold" value="1" step="0.1">
                    <span style="margin-left: 10px; font-size: 13px; color: #666;">（绝对差异小于此值标绿）</span>
                </div>
                <div class="form-row">
                    <label>输出文件名:</label>
                    <input type="text" id="aggOutputFile" value="聚合比对结果.xlsx">
                </div>
            </div>
            
            <div class="btn-row">
                <button class="btn-primary" onclick="runAggregateCompare()">🚀 开始聚合比对</button>
                <button class="btn-success" onclick="openAggResult()">打开结果</button>
                <button class="btn-secondary" onclick="openDir3()">打开目录</button>
            </div>
        </div>
        
        <!-- 运行日志（共享） -->
        <div class="section" style="margin-top: 20px;">
            <div class="section-title">运行日志</div>
            <div class="log-box" id="logBox">欢迎使用Excel比对工具!
[指标比对] 基于基准文件匹配横向数据
[指标+维度比对] 基于维度列匹配完整数据表

提示: 请直接输入文件的完整路径，或先点击"生成测试文件"</div>
        </div>
    </div>
    
    <script>
        // Tab切换
        function switchTab(tabNum) {
            document.querySelectorAll('.tab').forEach((t, i) => {
                t.classList.toggle('active', i === tabNum - 1);
            });
            document.querySelectorAll('.tab-content').forEach((t, i) => {
                t.classList.toggle('active', i === tabNum - 1);
            });
            const tabNames = ['[指标比对]', '[指标+维度比对]', '[聚合+比对]'];
            log('\\n切换到: ' + tabNames[tabNum - 1]);
        }
        
        function log(msg) {
            const box = document.getElementById('logBox');
            box.textContent += '\\n' + msg;
            box.scrollTop = box.scrollHeight;
        }
        
        function clearLog() {
            document.getElementById('logBox').textContent = '';
        }
        
        function showLoading() {
            document.getElementById('loadingOverlay').classList.add('show');
        }
        
        function hideLoading() {
            document.getElementById('loadingOverlay').classList.remove('show');
        }
        
        async function api(action, data) {
            try {
                const resp = await fetch('/api', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({action, ...data})
                });
                return await resp.json();
            } catch(e) {
                return {success: false, message: '请求失败: ' + e.message};
            }
        }
        
        // Tab 1 功能（保持不变）
        async function generateTest() {
            log('\\n[指标比对] 生成测试文件...');
            const workDir = document.getElementById('workDir').value;
            const result = await api('generate_test', {workDir});
            if (result.success) {
                log(result.message);
                document.getElementById('baseFile').value = result.baseFile;
                document.getElementById('dataAFile').value = result.dataAFile;
                document.getElementById('dataBFile').value = result.dataBFile;
                log('文件路径已自动填充!');
            } else {
                log('错误: ' + result.message);
            }
        }
        
        async function runCompare() {
            const data = {
                workDir: document.getElementById('workDir').value,
                baseFile: document.getElementById('baseFile').value,
                dataAFile: document.getElementById('dataAFile').value,
                dataBFile: document.getElementById('dataBFile').value,
                outputFile: document.getElementById('outputFile').value,
                decimalPlaces: parseInt(document.getElementById('decimalPlaces').value),
                greenTh: parseFloat(document.getElementById('greenTh').value)
            };
            
            if (!data.baseFile) { alert('请输入基准文件路径'); return; }
            if (!data.dataAFile) { alert('请输入输入1文件路径'); return; }
            if (!data.dataBFile) { alert('请输入输入2文件路径'); return; }
            
            log('\\n========================================');
            log('[指标比对] 开始对比...');
            log('小数位数: ' + data.decimalPlaces + ' 位');
            log('阈值: |差异%| < ' + data.greenTh + '% 或 A=B 为绿色');
            
            // 显示loading
            showLoading();
            
            try {
                const result = await api('compare', data);
                if (result.success) {
                    log(result.message);
                    alert('✅ 对比完成！');
                } else {
                    log('错误: ' + result.message);
                    alert('❌ 对比失败: ' + result.message);
                }
            } catch (error) {
                log('错误: ' + error.message);
                alert('❌ 对比异常: ' + error.message);
            } finally {
                // 隐藏loading
                hideLoading();
            }
        }
        
        async function openResult() {
            const workDir = document.getElementById('workDir').value;
            const outputFile = document.getElementById('outputFile').value;
            await api('open_file', {path: workDir + '/' + outputFile});
        }
        
        async function openDir() {
            const workDir = document.getElementById('workDir').value;
            await api('open_dir', {path: workDir});
        }
        
        async function browseFile(inputId) {
            log('正在打开文件选择对话框...');
            const workDir = document.getElementById('workDir').value;
            const result = await api('browse_file', {workDir});
            if (result.success && result.path) {
                document.getElementById(inputId).value = result.path;
                log('已选择: ' + result.path);
            } else if (result.message) {
                log(result.message);
            }
        }
        
        async function browseDir() {
            log('正在打开目录选择对话框...');
            const result = await api('browse_dir', {});
            if (result.success && result.path) {
                document.getElementById('workDir').value = result.path;
                log('工作目录: ' + result.path);
            }
        }
        
        // Tab 2 功能（新增）
        async function generateDimensionTest() {
            log('\\n[指标+维度比对] 生成测试文件...');
            const workDir = document.getElementById('workDir2').value;
            const result = await api('generate_dimension_test', {workDir});
            if (result.success) {
                log(result.message);
                document.getElementById('tableAFile').value = result.tableAFile;
                document.getElementById('tableBFile').value = result.tableBFile;
                log('文件路径已自动填充!');
            } else {
                log('错误: ' + result.message);
            }
        }
        
        async function runDimensionCompare() {
            const data = {
                workDir: document.getElementById('workDir2').value,
                tableAFile: document.getElementById('tableAFile').value,
                tableBFile: document.getElementById('tableBFile').value,
                keyColumns: parseInt(document.getElementById('keyColumns').value),
                diffThreshold: parseFloat(document.getElementById('diffThreshold').value),
                outputFile: document.getElementById('outputFile2').value
            };
            
            if (!data.tableAFile) { alert('请输入表A文件路径'); return; }
            if (!data.tableBFile) { alert('请输入表B文件路径'); return; }
            if (data.keyColumns < 1) { alert('基准列数量至少为1'); return; }
            
            log('\\n========================================');
            log('[指标+维度比对] 开始对比...');
            log('基准列数量: 前' + data.keyColumns + '列');
            log('差异阈值: ' + data.diffThreshold);
            log('匹配规则: 忽略空格、下划线、括号差异');
            
            // 显示loading
            showLoading();
            
            try {
                const result = await api('dimension_compare', data);
                if (result.success) {
                    log(result.message);
                    alert('✅ 对比完成！');
                } else {
                    log('错误: ' + result.message);
                    alert('❌ 对比失败: ' + result.message);
                }
            } catch (error) {
                log('错误: ' + error.message);
                alert('❌ 对比异常: ' + error.message);
            } finally {
                // 隐藏loading
                hideLoading();
            }
        }
        
        async function openDimensionResult() {
            const workDir = document.getElementById('workDir2').value;
            const outputFile = document.getElementById('outputFile2').value;
            await api('open_file', {path: workDir + '/' + outputFile});
        }
        
        async function openDir2() {
            const workDir = document.getElementById('workDir2').value;
            await api('open_dir', {path: workDir});
        }
        
        async function browseFile2(inputId) {
            log('正在打开文件选择对话框...');
            const workDir = document.getElementById('workDir2').value;
            const result = await api('browse_file', {workDir});
            if (result.success && result.path) {
                document.getElementById(inputId).value = result.path;
                log('已选择: ' + result.path);
            } else if (result.message) {
                log(result.message);
            }
        }
        
        async function browseDir2() {
            log('正在打开目录选择对话框...');
            const result = await api('browse_dir', {});
            if (result.success && result.path) {
                document.getElementById('workDir2').value = result.path;
                log('工作目录: ' + result.path);
            }
        }
        
        // ============ Tab3: 聚合+比对 功能 ============
        
        async function browseDir3() {
            log('正在打开目录选择对话框...');
            const result = await api('browse_dir', {});
            if (result.success && result.path) {
                document.getElementById('workDir3').value = result.path;
                log('工作目录: ' + result.path);
            }
        }
        
        async function parseHeaders() {
            const tableAFile = document.getElementById('aggTableAFile').value;
            const tableBFile = document.getElementById('aggTableBFile').value;
            
            if (!tableAFile) { alert('请选择表A文件'); return; }
            if (!tableBFile) { alert('请选择表B文件'); return; }
            
            log('\\n========================================');
            log('[聚合+比对] 解析表头...');
            
            const result = await api('parse_headers', { tableAFile, tableBFile });
            
            if (result.success) {
                log('成功解析表头，共 ' + result.headers.length + ' 列');
                log('列名: ' + result.headers.join(', '));
                
                // 显示列选择区域
                document.getElementById('columnSelectionSection').style.display = 'block';
                
                // 填充维度列和指标列的多选框
                const dimContainer = document.getElementById('dimensionColumns');
                const indContainer = document.getElementById('indicatorColumns');
                
                dimContainer.innerHTML = '';
                indContainer.innerHTML = '';
                
                result.headers.forEach(header => {
                    // 维度列选项
                    const dimOption = document.createElement('div');
                    dimOption.className = 'multiselect-option';
                    dimOption.innerHTML = `<input type="checkbox" value="${header}" onchange="updateDimCheckbox(this)"> ${header}`;
                    dimContainer.appendChild(dimOption);
                    
                    // 指标列选项
                    const indOption = document.createElement('div');
                    indOption.className = 'multiselect-option';
                    indOption.innerHTML = `<input type="checkbox" value="${header}" onchange="updateIndCheckbox(this)"> ${header}`;
                    indContainer.appendChild(indOption);
                });
                
                alert('✅ 表头解析完成！请选择维度列和指标列');
            } else {
                log('错误: ' + result.message);
                alert('❌ 解析失败: ' + result.message);
            }
        }
        
        function updateDimCheckbox(checkbox) {
            // 如果在维度列中选中，则在指标列中取消
            if (checkbox.checked) {
                const indCheckboxes = document.querySelectorAll('#indicatorColumns input[type="checkbox"]');
                indCheckboxes.forEach(cb => {
                    if (cb.value === checkbox.value) {
                        cb.checked = false;
                    }
                });
            }
        }
        
        function updateIndCheckbox(checkbox) {
            // 如果在指标列中选中，则在维度列中取消
            if (checkbox.checked) {
                const dimCheckboxes = document.querySelectorAll('#dimensionColumns input[type="checkbox"]');
                dimCheckboxes.forEach(cb => {
                    if (cb.value === checkbox.value) {
                        cb.checked = false;
                    }
                });
            }
        }
        
        function toggleAllDimensions() {
            const dimCheckboxes = document.querySelectorAll('#dimensionColumns input[type="checkbox"]');
            const indCheckboxes = document.querySelectorAll('#indicatorColumns input[type="checkbox"]');
            
            // 获取已在指标列选中的项
            const selectedInIndicators = new Set();
            indCheckboxes.forEach(cb => {
                if (cb.checked) {
                    selectedInIndicators.add(cb.value);
                }
            });
            
            // 检查当前可选的维度列是否全部选中
            let selectableCount = 0;
            let selectedCount = 0;
            dimCheckboxes.forEach(cb => {
                if (!selectedInIndicators.has(cb.value)) {
                    selectableCount++;
                    if (cb.checked) selectedCount++;
                }
            });
            
            const shouldSelect = selectedCount < selectableCount;
            
            // 切换选中状态（跳过已在指标列选中的项）
            dimCheckboxes.forEach(cb => {
                if (!selectedInIndicators.has(cb.value)) {
                    cb.checked = shouldSelect;
                }
            });
        }
        
        function toggleAllIndicators() {
            const dimCheckboxes = document.querySelectorAll('#dimensionColumns input[type="checkbox"]');
            const indCheckboxes = document.querySelectorAll('#indicatorColumns input[type="checkbox"]');
            
            // 获取已在维度列选中的项
            const selectedInDimensions = new Set();
            dimCheckboxes.forEach(cb => {
                if (cb.checked) {
                    selectedInDimensions.add(cb.value);
                }
            });
            
            // 检查当前可选的指标列是否全部选中
            let selectableCount = 0;
            let selectedCount = 0;
            indCheckboxes.forEach(cb => {
                if (!selectedInDimensions.has(cb.value)) {
                    selectableCount++;
                    if (cb.checked) selectedCount++;
                }
            });
            
            const shouldSelect = selectedCount < selectableCount;
            
            // 切换选中状态（跳过已在维度列选中的项）
            indCheckboxes.forEach(cb => {
                if (!selectedInDimensions.has(cb.value)) {
                    cb.checked = shouldSelect;
                }
            });
        }
        
        async function runAggregateCompare() {
            const tableAFile = document.getElementById('aggTableAFile').value;
            const tableBFile = document.getElementById('aggTableBFile').value;
            const workDir = document.getElementById('workDir3').value;
            const outputFile = document.getElementById('aggOutputFile').value;
            const diffThreshold = parseFloat(document.getElementById('aggDiffThreshold').value);
            
            // 获取选中的维度列和指标列
            const dimColumns = Array.from(document.querySelectorAll('#dimensionColumns input[type="checkbox"]:checked')).map(cb => cb.value);
            const indColumns = Array.from(document.querySelectorAll('#indicatorColumns input[type="checkbox"]:checked')).map(cb => cb.value);
            
            if (!tableAFile || !tableBFile) { alert('请先选择表A和表B文件'); return; }
            if (indColumns.length === 0) { alert('请至少选择一个指标列'); return; }
            
            const data = {
                workDir, tableAFile, tableBFile, outputFile,
                dimColumns, indColumns, diffThreshold
            };
            
            log('\\n========================================');
            log('[聚合+比对] 开始处理...');
            log('维度列: ' + dimColumns.join(', '));
            log('指标列: ' + indColumns.join(', '));
            log('差异阈值: ' + diffThreshold);
            
            showLoading();
            
            try {
                const result = await api('aggregate_compare', data);
                if (result.success) {
                    log(result.message);
                    alert('✅ 聚合比对完成！');
                } else {
                    log('错误: ' + result.message);
                    alert('❌ 处理失败: ' + result.message);
                }
            } catch (error) {
                log('错误: ' + error.message);
                alert('❌ 处理异常: ' + error.message);
            } finally {
                hideLoading();
            }
        }
        
        async function openAggResult() {
            const workDir = document.getElementById('workDir3').value;
            const outputFile = document.getElementById('aggOutputFile').value;
            const fullPath = workDir + '/' + outputFile;
            
            log('打开文件: ' + fullPath);
            const result = await api('open_file', { path: fullPath });
            if (!result.success) {
                alert('无法打开文件: ' + result.message);
            }
        }
        
        async function openDir3() {
            const workDir = document.getElementById('workDir3').value;
            log('打开目录: ' + workDir);
            const result = await api('open_dir', { path: workDir });
            if (!result.success) {
                alert('无法打开目录: ' + result.message);
            }
        }
    </script>
    
    <!-- Loading Overlay -->
    <div class="loading-overlay" id="loadingOverlay">
        <div class="loading-content">
            <div class="spinner"></div>
            <div class="loading-text">正在比对，请稍候...</div>
        </div>
    </div>
</body>
</html>
'''


class RequestHandler(BaseHTTPRequestHandler):
    """HTTP请求处理"""
    
    def log_message(self, format, *args):
        pass  # 禁用默认日志
    
    def _convert_xls_to_xlsx(self, xls_path):
        """将.xls文件转换为临时.xlsx文件
        
        Args:
            xls_path: .xls文件路径
            
        Returns:
            临时.xlsx文件路径，如果转换失败则返回None
        """
        if not XLRD_OK:
            raise Exception('缺少xlrd库，无法读取.xls文件。请安装xlrd或将文件转换为.xlsx格式')
        
        try:
            # 使用xlrd读取.xls文件
            xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
            xls_sheet = xls_book.sheet_by_index(0)
            
            # 创建临时.xlsx文件
            import tempfile
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            
            # 使用openpyxl写入.xlsx
            wb = Workbook()
            ws = wb.active
            
            # 复制数据
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    # 处理不同的单元格类型
                    if xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                        # 日期类型需要特殊处理
                        from datetime import datetime
                        cell_value = xlrd.xldate_as_datetime(cell_value, xls_book.datemode)
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            
            wb.save(temp_path)
            wb.close()
            return temp_path
            
        except Exception as e:
            raise Exception(f'转换.xls文件失败: {str(e)}')
    
    def _load_workbook_safe(self, file_path, data_only=True):
        """安全加载workbook，自动处理.xls格式
        
        Args:
            file_path: Excel文件路径
            data_only: 是否只读取数据值（不读取公式）
            
        Returns:
            (workbook对象, 临时文件路径或None)
        """
        temp_file = None
        
        # 检查文件扩展名
        _, ext = os.path.splitext(file_path.lower())
        
        if ext == '.xls':
            # 转换.xls为临时.xlsx
            temp_file = self._convert_xls_to_xlsx(file_path)
            file_path = temp_file
        elif ext != '.xlsx' and ext != '.xlsm':
            raise Exception(f'不支持的文件格式: {ext}。请使用.xlsx, .xlsm或.xls格式')
        
        # 加载workbook
        wb = load_workbook(file_path, data_only=data_only)
        
        return wb, temp_file
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.end_headers()
        self.wfile.write(HTML_TEMPLATE.encode('utf-8'))
    
    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8')
        
        try:
            data = json.loads(body)
            action = data.get('action', '')
            
            if action == 'generate_test':
                result = self.generate_test(data.get('workDir', WORK_DIR))
            elif action == 'compare':
                result = self.run_compare(data)
            elif action == 'generate_dimension_test':
                result = self.generate_dimension_test(data.get('workDir', WORK_DIR))
            elif action == 'dimension_compare':
                result = self.run_dimension_compare(data)
            elif action == 'parse_headers':
                result = self.parse_table_headers(data)
            elif action == 'aggregate_compare':
                result = self.run_aggregate_compare(data)
            elif action == 'open_file':
                result = self.open_file(data.get('path', ''))
            elif action == 'open_dir':
                result = self.open_dir(data.get('path', ''))
            elif action == 'browse_file':
                result = self.browse_file_dialog(data.get('workDir', WORK_DIR))
            elif action == 'browse_dir':
                result = self.browse_dir_dialog()
            else:
                result = {'success': False, 'message': '未知操作'}
                
        except Exception as e:
            result = {'success': False, 'message': str(e)}
        
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(result, ensure_ascii=False).encode('utf-8'))
    
    def generate_test(self, workdir):
        """生成测试文件"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        if not os.path.exists(workdir):
            return {'success': False, 'message': '工作目录不存在'}
        
        try:
            # 基准文件
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="指标名称")
            indicators = [
                "正常_完全相同", "正常_小差异_0.5%", "正常_临界_1%", "正常_中等_5%",
                "正常_较大_50%", "正常_超大_150%", "特殊_B为零", "特殊_负数",
                "缺失_A无数据", "缺失_B无数据", "缺失_都无数据"
            ]
            for i, name in enumerate(indicators, 2):
                ws.cell(row=i, column=1, value=name)
            base_path = os.path.join(workdir, "test_base.xlsx")
            wb.save(base_path)
            
            # 数据A
            wb = Workbook()
            ws = wb.active
            data_a = [
                ("正常_完全相同", 1000000), ("正常_小差异_0.5%", 1005000),
                ("正常_临界_1%", 1010000), ("正常_中等_5%", 1050000),
                ("正常_较大_50%", 1500000), ("正常_超大_150%", 2500000),
                ("特殊_B为零", 100), ("特殊_负数", -500),
                ("缺失_A无数据", None), ("缺失_都无数据", None)
            ]
            for col, (h, v) in enumerate(data_a, 1):
                ws.cell(row=1, column=col, value=h)
                ws.cell(row=2, column=col, value=v)
            data_a_path = os.path.join(workdir, "test_data_a.xlsx")
            wb.save(data_a_path)
            
            # 数据B
            wb = Workbook()
            ws = wb.active
            data_b = [
                ("正常_完全相同", 1000000), ("正常_小差异_0.5%", 1000000),
                ("正常_临界_1%", 1000000), ("正常_中等_5%", 1000000),
                ("正常_较大_50%", 1000000), ("正常_超大_150%", 1000000),
                ("特殊_B为零", 0), ("特殊_负数", -400),
                ("缺失_B无数据", None), ("缺失_都无数据", None)
            ]
            for col, (h, v) in enumerate(data_b, 1):
                ws.cell(row=1, column=col, value=h)
                ws.cell(row=2, column=col, value=v)
            data_b_path = os.path.join(workdir, "test_data_b.xlsx")
            wb.save(data_b_path)
            
            return {
                'success': True,
                'message': '测试文件已生成:\n  - test_base.xlsx\n  - test_data_a.xlsx\n  - test_data_b.xlsx',
                'baseFile': base_path,
                'dataAFile': data_a_path,
                'dataBFile': data_b_path
            }
            
        except Exception as e:
            return {'success': False, 'message': str(e)}
    
    def generate_dimension_test(self, workdir):
        """生成维度比对测试文件"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        if not os.path.exists(workdir):
            return {'success': False, 'message': '工作目录不存在'}
        
        try:
            # 表A：包含维度列和指标列
            wb = Workbook()
            ws = wb.active
            
            # 表头
            headers_a = ['险种', '渠道', '指标1', '指标2', '指标3']
            for col, h in enumerate(headers_a, 1):
                ws.cell(row=1, column=col, value=h)
            
            # 数据行
            data_a = [
                ['车险', '银行', 1000, 2000, 3000],
                ['车_险', '电销', 1100, 2100, 3100],  # 维度键带下划线
                ['健康险', '代理', 1200, 2200, 3200],
                ['意外险（短期）', '网销', 1300, 2300, 3300],  # 维度键带括号
                ['寿险 A', '银行', 1400, 2400, 3400],  # 维度键带空格
                ['财产险', '直销', 1500, 2500, 3500],  # A独有
            ]
            
            for row_idx, row_data in enumerate(data_a, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            table_a_path = os.path.join(workdir, "test_table_a.xlsx")
            wb.save(table_a_path)
            
            # 表B：包含维度列和指标列（部分不同）
            wb = Workbook()
            ws = wb.active
            
            # 表头（指标2不在B表中，会被过滤）
            headers_b = ['险种', '渠道', '指标1', '指标3', '指标4']
            for col, h in enumerate(headers_b, 1):
                ws.cell(row=1, column=col, value=h)
            
            # 数据行
            data_b = [
                ['车险', '银行', 1000, 3000, 4000],  # 完全匹配
                ['车险', '电销', 1100, 3100, 4100],  # 能匹配（忽略下划线）
                ['健康险', '代理', 1250, 3250, 4250],  # 数据不同
                ['意外险【短期】', '网销', 1300, 3300, 4300],  # 能匹配（忽略括号）
                ['寿险A', '银行', 1400, 3400, 4400],  # 能匹配（忽略空格）
                ['重疾险', '网销', 1600, 3600, 4600],  # B独有
            ]
            
            for row_idx, row_data in enumerate(data_b, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            table_b_path = os.path.join(workdir, "test_table_b.xlsx")
            wb.save(table_b_path)
            
            return {
                'success': True,
                'message': '维度比对测试文件已生成:\n  - test_table_a.xlsx (表A)\n  - test_table_b.xlsx (表B)\n\n说明:\n  - 前2列为维度列（险种、渠道）\n  - 包含完全匹配、模糊匹配、不匹配的行\n  - 表B的指标列为基准',
                'tableAFile': table_a_path,
                'tableBFile': table_b_path
            }
            
        except Exception as e:
            import traceback
            return {'success': False, 'message': str(e) + '\n' + traceback.format_exc()}
    
    def run_dimension_compare(self, data):
        """运行维度比对"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        try:
            workdir = data.get('workDir', WORK_DIR)
            table_a_file = data.get('tableAFile', '')
            table_b_file = data.get('tableBFile', '')
            key_columns = int(data.get('keyColumns', 1))
            diff_threshold = float(data.get('diffThreshold', 1))
            output_file = data.get('outputFile', 'dimension_compare_result.xlsx')
            
            # 读取表A和表B
            table_a = self._read_full_table(table_a_file)
            table_b = self._read_full_table(table_b_file)
            
            # 提取文件名（用于error标记）
            table_a_name = os.path.basename(table_a_file).replace('.xlsx', '').replace('.xls', '')
            table_b_name = os.path.basename(table_b_file).replace('.xlsx', '').replace('.xls', '')
            
            # 生成结果
            output_path = os.path.join(workdir, output_file)
            self._create_dimension_result(
                output_path, table_a, table_b, key_columns,
                table_a_name, table_b_name, diff_threshold,
                table_a_file, table_b_file
            )
            
            return {
                'success': True,
                'message': '维度比对完成!\n表A: {} 行\n表B: {} 行\n基准列: 前{}列\n差异阈值: {}\n结果已保存: {}'.format(
                    len(table_a['data']), len(table_b['data']), key_columns, diff_threshold, output_file
                )
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'message': str(e)}
    
    def parse_table_headers(self, data):
        """解析表A和表B的表头，去重后返回"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        try:
            table_a_file = data.get('tableAFile', '')
            table_b_file = data.get('tableBFile', '')
            
            if not table_a_file or not table_b_file:
                return {'success': False, 'message': '请提供表A和表B文件路径'}
            
            # 读取两个表的表头
            table_a = self._read_full_table(table_a_file)
            table_b = self._read_full_table(table_b_file)
            
            # 合并表头并去重（保持顺序，忽略下划线、空格和括号的差异）
            all_headers = []
            seen = set()  # 存储标准化后的字符串
            for header in table_a['headers'] + table_b['headers']:
                normalized = self._normalize_string(header)
                if normalized not in seen:
                    all_headers.append(header)  # 保存原始列名
                    seen.add(normalized)
            
            return {
                'success': True,
                'headers': all_headers,
                'message': f'成功解析表头，共{len(all_headers)}列'
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'message': str(e)}
    
    def run_aggregate_compare(self, data):
        """运行聚合比对：对A和B分别聚合，然后比对"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        try:
            workdir = data.get('workDir', WORK_DIR)
            table_a_file = data.get('tableAFile', '')
            table_b_file = data.get('tableBFile', '')
            dim_columns = data.get('dimColumns', [])
            ind_columns = data.get('indColumns', [])
            diff_threshold = float(data.get('diffThreshold', 1))
            output_file = data.get('outputFile', '聚合比对结果.xlsx')
            
            if not ind_columns:
                return {'success': False, 'message': '请至少选择一个指标列'}
            
            # 读取原始表
            table_a_raw = self._read_full_table(table_a_file)
            table_b_raw = self._read_full_table(table_b_file)
            
            # 提取文件名
            table_a_name = os.path.basename(table_a_file).replace('.xlsx', '').replace('.xls', '')
            table_b_name = os.path.basename(table_b_file).replace('.xlsx', '').replace('.xls', '')
            
            # 聚合A和B
            agg_a = self._aggregate_table(table_a_raw, dim_columns, ind_columns, table_a_name)
            agg_b = self._aggregate_table(table_b_raw, dim_columns, ind_columns, table_b_name)
            
            # 生成结果文件（5个sheet）
            output_path = os.path.join(workdir, output_file)
            self._create_aggregate_result(
                output_path, 
                agg_a, agg_b,
                table_a_file, table_b_file,
                table_a_name, table_b_name,
                len(dim_columns), diff_threshold
            )
            
            return {
                'success': True,
                'message': f'''聚合比对完成！
表A聚合后: {len(agg_a['data'])} 行
表B聚合后: {len(agg_b['data'])} 行
维度列: {len(dim_columns)} 个
指标列: {len(ind_columns)} 个
结果已保存: {output_file}'''
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'message': str(e)}
    
    def run_compare(self, data):
        """运行对比"""
        if not OPENPYXL_OK:
            return {'success': False, 'message': '缺少openpyxl库'}
        
        try:
            workdir = data.get('workDir', WORK_DIR)
            base_file = data.get('baseFile', '')
            data_a_file = data.get('dataAFile', '')
            data_b_file = data.get('dataBFile', '')
            output_file = data.get('outputFile', 'compare_result.xlsx')
            decimal_places = int(data.get('decimalPlaces', 6))
            green_th = float(data.get('greenTh', 1.0))
            
            # 读取基准
            base_names = self._read_base(base_file)
            
            # 读取数据
            data_a = self._read_horizontal(data_a_file)
            data_b = self._read_horizontal(data_b_file)
            
            # 提取文件名（用于表头显示）
            data_a_name = os.path.basename(data_a_file).replace('.xlsx', '').replace('.xls', '')
            data_b_name = os.path.basename(data_b_file).replace('.xlsx', '').replace('.xls', '')
            
            # 生成结果
            output_path = os.path.join(workdir, output_file)
            self._create_result(output_path, base_names, data_a, data_b, decimal_places, green_th, 
                              data_a_name, data_b_name, base_file, data_a_file, data_b_file)
            
            return {
                'success': True, 
                'message': '基准: {} 个指标\n输入1: {} 个数据\n输入2: {} 个数据\n小数位数: {} 位\n========================================\n[完成] 结果已保存: {}'.format(
                    len(base_names), len(data_a), len(data_b), decimal_places, output_file
                )
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'message': str(e)}
    
    def _read_base(self, path):
        # 处理中文路径
        if sys.platform == 'win32' and isinstance(path, str):
            # Windows上确保路径是Unicode字符串
            path = os.path.normpath(path)
        
        wb, temp_file = self._load_workbook_safe(path, data_only=True)
        try:
            ws = wb.active
            names = []
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=1).value
                if v:
                    names.append(str(v).strip())
            return names
        finally:
            wb.close()
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
    
    def _read_horizontal(self, path):
        # 处理中文路径
        if sys.platform == 'win32' and isinstance(path, str):
            # Windows上确保路径是Unicode字符串
            path = os.path.normpath(path)
        
        wb, temp_file = self._load_workbook_safe(path, data_only=True)
        try:
            ws = wb.active
            data = {}
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                if h:
                    # 保存原始key和标准化key的映射
                    original_key = str(h).strip()
                    data[original_key] = ws.cell(row=2, column=col).value
            return data
        finally:
            wb.close()
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
    
    def _normalize_key(self, key):
        """标准化指标名称，只忽略下划线"""
        return key.replace('_', '').lower()
    
    def _find_value(self, data_dict, target_key):
        """根据标准化规则查找值，忽略下划线差异"""
        # 先尝试精确匹配
        if target_key in data_dict:
            return data_dict[target_key]
        
        # 标准化后模糊匹配（只忽略下划线）
        normalized_target = self._normalize_key(target_key)
        for key, value in data_dict.items():
            if self._normalize_key(key) == normalized_target:
                return value
        
        return None
    
    def _parse_num(self, v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        s = str(v).strip().replace(',', '').replace(' ', '')
        if not s or s.lower() in ['error', '#value!', 'none', 'null']:
            return None
        try:
            return Decimal(s)
        except:
            return None
    
    def _create_result(self, output, names, data_a, data_b, decimal_places, green_th, 
                      data_a_name='A', data_b_name='B', base_file=None, data_a_file=None, data_b_file=None):
        GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        HEADER = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        LEGEND_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "比对结果"
        
        # 构造格式化字符串（根据小数位数）
        format_str = '0.' + '0' * decimal_places
        
        # 表头（第1行）- 使用实际文件名
        headers = ["指标名称", data_a_name, data_b_name, 
                  f"差额({data_a_name}-{data_b_name})", "差异%"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HEADER
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = border
        
        # 图例放在右上角 G1:H2（与表头同行及下一行）
        legend_col = 7  # G列
        cell_g1 = ws.cell(row=1, column=legend_col, 
                         value="{}={} 或 |差异%|<{}%".format(data_a_name, data_b_name, green_th))
        cell_g1.border = border
        cell_g1.fill = LEGEND_FILL
        cell_g1.alignment = Alignment(horizontal='left')
        cell_g1.font = Font(size=10)
        
        cell_h1 = ws.cell(row=1, column=legend_col+1, value="绿色")
        cell_h1.fill = GREEN
        cell_h1.border = border
        cell_h1.alignment = Alignment(horizontal='center')
        cell_h1.font = Font(size=10)
        
        cell_g2 = ws.cell(row=2, column=legend_col, value="其他情况")
        cell_g2.border = border
        cell_g2.fill = LEGEND_FILL
        cell_g2.alignment = Alignment(horizontal='left')
        cell_g2.font = Font(size=10)
        
        cell_h2 = ws.cell(row=2, column=legend_col+1, value="红色")
        cell_h2.fill = RED
        cell_h2.border = border
        cell_h2.alignment = Alignment(horizontal='center')
        cell_h2.font = Font(size=10)
        
        # 数据行（从第2行开始）
        current_row = 2
        for name in names:
            ws.cell(row=current_row, column=1, value=name).border = border
            
            # 使用模糊匹配查找值
            va = self._find_value(data_a, name)
            vb = self._find_value(data_b, name)
            pa = self._parse_num(va)
            pb = self._parse_num(vb)
            
            # A列
            if pa is not None:
                ws.cell(row=current_row, column=2, value=float(pa)).border = border
            else:
                ws.cell(row=current_row, column=2, value="error").border = border
                
            # B列
            if pb is not None:
                ws.cell(row=current_row, column=3, value=float(pb)).border = border
            else:
                ws.cell(row=current_row, column=3, value="error").border = border
                
            # 差额 (A-B)
            if pa is not None and pb is not None:
                diff = pa - pb
                diff_formatted = float(diff.quantize(Decimal(format_str), rounding=ROUND_HALF_UP))
                ws.cell(row=current_row, column=4, value=diff_formatted).border = border
            else:
                ws.cell(row=current_row, column=4, value="#VALUE!").border = border
                diff = None
                
            # 差异百分比 (A-B)/A * 100
            cell = ws.cell(row=current_row, column=5)
            cell.border = border
            
            if pa is not None and pb is not None:
                # 判断A和B是否相等
                if pa == pb:
                    cell.value = "0%"
                    cell.fill = GREEN  # A=B 时为绿色
                elif pa == 0:
                    # A为0时，无法计算百分比
                    cell.value = "#VALUE!"
                    cell.fill = RED
                else:
                    # 计算百分比: (A-B)/A * 100
                    pct = (diff / pa) * 100
                    pct_formatted = float(pct.quantize(Decimal(format_str), rounding=ROUND_HALF_UP))
                    cell.value = "{}%".format(pct_formatted)
                    
                    # 颜色判断：|差异%| < green_th 为绿色，否则为红色
                    abs_pct = abs(pct)
                    if abs_pct < green_th:
                        cell.fill = GREEN
                    else:
                        cell.fill = RED
            else:
                cell.value = "#VALUE!"
                cell.fill = RED
            
            current_row += 1
                
        # 调整列宽
        for col, w in enumerate([22, 18, 18, 16, 16, 16, 10], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        
        # 冻结表头（第1行）
        ws.freeze_panes = 'A2'
        
        # 复制源文件到结果workbook
        if base_file and os.path.exists(base_file):
            self._copy_sheet_from_file(wb, base_file, "基准文件")
        if data_a_file and os.path.exists(data_a_file):
            self._copy_sheet_from_file(wb, data_a_file, f"源文件_{data_a_name}")
        if data_b_file and os.path.exists(data_b_file):
            self._copy_sheet_from_file(wb, data_b_file, f"源文件_{data_b_name}")
        
        # 保存文件，处理中文路径编码
        try:
            wb.save(output)
        except Exception as e:
            # 如果保存失败，尝试用不同的编码
            if sys.platform == 'win32':
                # Windows上尝试使用UTF-8
                output_bytes = output.encode('utf-8')
                wb.save(output_bytes.decode('utf-8'))
            else:
                raise e
    
    def open_file(self, path):
        if os.path.exists(path):
            if sys.platform == 'darwin':
                os.system('open "{}"'.format(path))
            elif sys.platform == 'win32':
                os.system('start "" "{}"'.format(path))
            else:
                os.system('xdg-open "{}"'.format(path))
            return {'success': True}
        return {'success': False, 'message': '文件不存在'}
    
    def open_dir(self, path):
        if os.path.exists(path):
            if sys.platform == 'darwin':
                os.system('open "{}"'.format(path))
            elif sys.platform == 'win32':
                os.system('explorer "{}"'.format(path))
            else:
                os.system('xdg-open "{}"'.format(path))
            return {'success': True}
        return {'success': False, 'message': '目录不存在'}
    
    def browse_file_dialog(self, initial_dir):
        """打开文件选择对话框"""
        try:
            if sys.platform == 'darwin':
                # macOS: 使用osascript
                script = '''
                tell application "System Events"
                    activate
                    set theFile to choose file with prompt "选择Excel文件" of type {"xlsx", "xls"}
                    return POSIX path of theFile
                end tell
                '''
                result = subprocess.run(['osascript', '-e', script], 
                                        capture_output=True, text=True, timeout=60)
                if result.returncode == 0 and result.stdout.strip():
                    return {'success': True, 'path': result.stdout.strip()}
                return {'success': False, 'message': '未选择文件'}
            else:
                # Windows/Linux: 直接使用tkinter（集成方式，解决PyInstaller subprocess循环问题）
                try:
                    import tkinter as tk
                    from tkinter import filedialog
                except ImportError:
                    return {'success': False, 'message': 'tkinter未安装'}
                
                # 确保初始目录存在
                if not initial_dir or not os.path.exists(initial_dir):
                    initial_dir = os.getcwd()
                
                # 创建隐藏的根窗口
                root = tk.Tk()
                root.withdraw()  # 隐藏主窗口
                
                # Windows上设置窗口置顶
                if sys.platform == 'win32':
                    try:
                        root.wm_attributes('-topmost', True)
                        root.focus_force()
                    except:
                        pass
                
                # 打开文件选择对话框
                file_path = filedialog.askopenfilename(
                    title='选择Excel文件',
                    initialdir=initial_dir,
                    filetypes=[
                        ('Excel文件', '*.xlsx *.xls'),
                        ('所有文件', '*.*')
                    ]
                )
                
                # 销毁根窗口
                root.destroy()
                
                if file_path:
                    return {'success': True, 'path': file_path}
                else:
                    return {'success': False, 'message': '未选择文件'}
        except subprocess.TimeoutExpired:
            return {'success': False, 'message': '选择超时'}
        except Exception as e:
            import traceback
            return {'success': False, 'message': str(e) + '\n' + traceback.format_exc()}
    
    def browse_dir_dialog(self):
        """打开目录选择对话框"""
        try:
            if sys.platform == 'darwin':
                # macOS: 使用osascript
                script = '''
                tell application "System Events"
                    activate
                    set theFolder to choose folder with prompt "选择工作目录"
                    return POSIX path of theFolder
                end tell
                '''
                result = subprocess.run(['osascript', '-e', script],
                                        capture_output=True, text=True, timeout=60)
                if result.returncode == 0 and result.stdout.strip():
                    return {'success': True, 'path': result.stdout.strip().rstrip('/')}
                return {'success': False, 'message': '未选择目录'}
            else:
                # Windows/Linux: 直接使用tkinter（集成方式，解决PyInstaller subprocess循环问题）
                try:
                    import tkinter as tk
                    from tkinter import filedialog
                except ImportError:
                    return {'success': False, 'message': 'tkinter未安装'}
                
                # 创建隐藏的根窗口
                root = tk.Tk()
                root.withdraw()  # 隐藏主窗口
                
                # Windows上设置窗口置顶
                if sys.platform == 'win32':
                    try:
                        root.wm_attributes('-topmost', True)
                        root.focus_force()
                    except:
                        pass
                
                # 打开目录选择对话框
                dir_path = filedialog.askdirectory(
                    title='选择工作目录',
                    initialdir=os.getcwd()
                )
                
                # 销毁根窗口
                root.destroy()
                
                if dir_path:
                    return {'success': True, 'path': dir_path}
                else:
                    return {'success': False, 'message': '未选择目录'}
        except subprocess.TimeoutExpired:
            return {'success': False, 'message': '选择超时'}
        except Exception as e:
            import traceback
            return {'success': False, 'message': str(e) + '\n' + traceback.format_exc()}
    
    def _read_full_table(self, file_path):
        """读取完整的Excel表格"""
        wb, temp_file = self._load_workbook_safe(file_path, data_only=True)
        try:
            ws = wb.active
            
            # 读取所有数据
            data = []
            headers = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    # 表头
                    headers = [str(cell) if cell is not None else f'列{i}' for i, cell in enumerate(row, 1)]
                else:
                    # 数据行（跳过全空行）
                    if any(cell is not None and str(cell).strip() != '' for cell in row):
                        data.append(list(row))
            
            return {
                'headers': headers,
                'data': data
            }
        finally:
            wb.close()
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
    
    def _copy_sheet_from_file(self, target_wb, source_file, sheet_name, highlight_rows=None):
        """从源文件复制sheet到目标workbook，可选择高亮指定行
        
        Args:
            target_wb: 目标workbook
            source_file: 源文件路径
            sheet_name: 新sheet名称
            highlight_rows: 需要标红的行号列表（从1开始，包含表头）
        """
        temp_file = None
        try:
            source_wb, temp_file = self._load_workbook_safe(source_file, data_only=True)
            source_ws = source_wb.active
            
            # 创建新sheet
            target_ws = target_wb.create_sheet(title=sheet_name)
            
            # 红色填充（用于标识不匹配的行）
            HIGHLIGHT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 复制数据
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    
                    # 复制格式
                    if cell.has_style:
                        try:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()
                        except:
                            pass
                    
                    # 如果该行需要高亮，覆盖背景色
                    if highlight_rows and cell.row in highlight_rows:
                        target_cell.fill = HIGHLIGHT_FILL
            
            # 复制列宽
            for col_letter in source_ws.column_dimensions:
                if col_letter in source_ws.column_dimensions:
                    target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
            
            # 复制行高
            for row_num in source_ws.row_dimensions:
                if row_num in source_ws.row_dimensions:
                    target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height
            
            # 冻结表头
            target_ws.freeze_panes = 'A2'
                    
        except Exception as e:
            print(f"复制sheet失败: {e}")
        finally:
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
    
    def _normalize_string(self, s):
        """
        标准化字符串，忽略：
        - 空格
        - 下划线 _
        - 中文括号 （）【】
        - 英文括号 ()[]
        """
        import re
        if s is None:
            return ''
        s = str(s).strip()
        # 移除空格
        s = s.replace(' ', '')
        # 移除下划线
        s = s.replace('_', '')
        # 移除各种括号
        s = re.sub(r'[()（）\[\]【】]', '', s)
        return s.lower()
    
    def _normalize_dimension_key(self, key_values):
        """
        标准化维度键，忽略：
        - 空格
        - 下划线 _
        - 中文括号 （）【】
        - 英文括号 ()[]
        """
        normalized = []
        for val in key_values:
            normalized.append(self._normalize_string(val))
        return tuple(normalized)
    
    def _create_dimension_result(self, output, table_a, table_b, key_columns, 
                                 table_a_name, table_b_name, diff_threshold,
                                 table_a_file=None, table_b_file=None):
        """生成维度比对结果Excel"""
        HEADER = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        ERROR_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        RED_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        ROW_MISSING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 行不匹配的红色标识
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "维度比对结果"
        
        headers_a = table_a['headers']
        headers_b = table_b['headers']
        data_a = table_a['data']
        data_b = table_b['data']
        
        # 1. 确定维度列和指标列
        dim_headers = headers_b[:key_columns]  # 维度列使用B表的表头
        indicators_a = headers_a[key_columns:]  # A表的指标列
        indicators_b = headers_b[key_columns:]  # B表的指标列
        
        # 2. 构建表头（维度列 + 指标列，指标列显示差异值）
        result_headers = list(dim_headers) + list(indicators_b)
        
        # 写入表头
        for col, h in enumerate(result_headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HEADER
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = border
        
        # 3. 构建A和B的索引（标准化键 -> (行数据, 原始行号)）
        a_index = {}
        a_row_nums = {}  # 标准化键 -> 源文件行号（从2开始，1是表头）
        for idx, row_data in enumerate(data_a):
            key_vals = row_data[:key_columns]
            norm_key = self._normalize_dimension_key(key_vals)
            a_index[norm_key] = row_data
            a_row_nums[norm_key] = idx + 2  # +2 因为: data_a是从0开始，源文件第1行是表头
        
        b_index = {}
        b_row_nums = {}  # 标准化键 -> 源文件行号
        b_keys_order = []  # 保持B表的行顺序
        for idx, row_data in enumerate(data_b):
            key_vals = row_data[:key_columns]
            norm_key = self._normalize_dimension_key(key_vals)
            b_index[norm_key] = row_data
            b_row_nums[norm_key] = idx + 2
            b_keys_order.append((norm_key, row_data[:key_columns]))
        
        # 4. 生成结果行
        result_rows = []
        matched_a_keys = set()
        unmatched_a_rows = set()  # A表中不匹配的行号
        unmatched_b_rows = set()  # B表中不匹配的行号
        
        # 遍历B表的行
        for norm_key, original_key_vals in b_keys_order:
            result_row = []
            result_row_meta = []  # 存储元数据：类型（diff/error_a/error_b）和原始值
            row_type = 'both'  # 记录行类型：'both'（都有）、'only_a'（只在A）、'only_b'（只在B）
            
            # 维度列（来自B表）
            for val in original_key_vals:
                result_row.append(val)
                result_row_meta.append(('dim', None))
            
            # 查找A表中是否有匹配的行
            if norm_key in a_index:
                # A和B都有
                matched_a_keys.add(norm_key)
                a_row = a_index[norm_key]
                b_row = b_index[norm_key]
                row_type = 'both'
                
                # 填充指标列（显示差异值 A - B）
                for ind in indicators_b:
                    if ind in indicators_a:
                        # A和B都有这个指标
                        a_idx = headers_a.index(ind)
                        b_idx = headers_b.index(ind)
                        a_val = a_row[a_idx] if a_idx < len(a_row) else None
                        b_val = b_row[b_idx] if b_idx < len(b_row) else None
                        
                        # 尝试计算差异
                        diff_val = self._calculate_diff(a_val, b_val, table_a_name, table_b_name)
                        result_row.append(diff_val)
                        result_row_meta.append(('diff', diff_val))
                    else:
                        # B有但A没有的指标
                        result_row.append(f'{table_a_name}表error')
                        result_row_meta.append(('error', None))
            else:
                # 只有B有，A没有
                row_type = 'only_b'
                unmatched_b_rows.add(b_row_nums[norm_key])  # 记录B表中不匹配的行号
                for ind in indicators_b:
                    result_row.append(f'{table_a_name}表error')
                    result_row_meta.append(('error', None))
            
            result_rows.append((result_row, result_row_meta, row_type))
        
        # 5. 添加A表独有的行
        for norm_key, a_row in a_index.items():
            if norm_key not in matched_a_keys:
                # 只有A有，B没有
                result_row = []
                result_row_meta = []
                row_type = 'only_a'
                unmatched_a_rows.add(a_row_nums[norm_key])  # 记录A表中不匹配的行号
                
                # 维度列（来自A表）
                original_key_vals = a_row[:key_columns]
                for val in original_key_vals:
                    result_row.append(val)
                    result_row_meta.append(('dim', None))
                
                # 指标列
                for ind in indicators_b:
                    if ind in indicators_a:
                        # A和B都有这个指标列，但这一行只在A表
                        result_row.append(f'{table_b_name}表error')
                        result_row_meta.append(('error', None))
                    else:
                        # 这个指标列只在B表，这一行也只在A表
                        # 显示两个error
                        result_row.append(f'{table_a_name}表error, {table_b_name}表error')
                        result_row_meta.append(('error', None))
                
                result_rows.append((result_row, result_row_meta, row_type))
        
        # 6. 写入数据行，并根据差异值标记颜色
        for row_idx, (row_data, row_meta, row_type) in enumerate(result_rows, 2):
            for col_idx, (value, meta) in enumerate(zip(row_data, row_meta), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # 根据单元格类型标记颜色
                if meta[0] == 'error':
                    # Error标记：红色背景
                    cell.fill = ERROR_FILL
                    cell.font = Font(color="FF0000")
                elif meta[0] == 'diff' and isinstance(meta[1], (int, float)):
                    # 差异值：根据阈值标记颜色
                    abs_diff = abs(meta[1])
                    if abs_diff < diff_threshold:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL
        
        # 7. 添加图例（放在右上角）
        legend_start_col = len(result_headers) + 2
        legend_row = 1
        
        # 图例标题
        legend_title = ws.cell(row=legend_row, column=legend_start_col, value="图例")
        legend_title.font = Font(bold=True)
        legend_title.border = border
        
        # 绿色图例
        legend_row += 1
        green_cell = ws.cell(row=legend_row, column=legend_start_col, value=f"|差异| < {diff_threshold}")
        green_cell.fill = GREEN_FILL
        green_cell.border = border
        
        # 红色图例
        legend_row += 1
        red_cell = ws.cell(row=legend_row, column=legend_start_col, value=f"|差异| ≥ {diff_threshold}")
        red_cell.fill = RED_FILL
        red_cell.border = border
        
        # 源文件行不匹配说明
        legend_row += 1
        missing_cell = ws.cell(row=legend_row, column=legend_start_col, value="不匹配行已在源文件sheet中标红")
        missing_cell.fill = ROW_MISSING_FILL
        missing_cell.border = border
        
        # 8. 调整列宽
        for col_idx, header in enumerate(result_headers, 1):
            col_letter = get_column_letter(col_idx)
            if col_idx <= key_columns:
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 16
        
        # 图例列宽
        legend_col_letter = get_column_letter(legend_start_col)
        ws.column_dimensions[legend_col_letter].width = 20
        
        # 冻结表头
        ws.freeze_panes = 'A2'
        
        # 复制源文件到结果workbook，并标红不匹配的行
        if table_a_file and os.path.exists(table_a_file):
            self._copy_sheet_from_file(wb, table_a_file, f"源文件_{table_a_name}", 
                                      highlight_rows=unmatched_a_rows if unmatched_a_rows else None)
        if table_b_file and os.path.exists(table_b_file):
            self._copy_sheet_from_file(wb, table_b_file, f"源文件_{table_b_name}",
                                      highlight_rows=unmatched_b_rows if unmatched_b_rows else None)
        
        # 9. 保存文件
        try:
            wb.save(output)
        except Exception as e:
            if sys.platform == 'win32':
                output_bytes = output.encode('utf-8')
                wb.save(output_bytes.decode('utf-8'))
            else:
                raise e
    
    def _calculate_diff(self, a_val, b_val, table_a_name, table_b_name):
        """计算差异值 B - A"""
        # 如果任一值为空，返回error
        if a_val is None or str(a_val).strip() == '':
            return f'{table_a_name}表error'
        if b_val is None or str(b_val).strip() == '':
            return f'{table_b_name}表error'
        
        # 尝试转换为数值
        try:
            a_num = float(a_val)
            b_num = float(b_val)
            return b_num - a_num  # 修改为 B - A
        except (ValueError, TypeError):
            # 无法转换为数值，返回error
            return f'无法计算差异'
    
    def _aggregate_table(self, table_data, dim_columns, ind_columns, table_name):
        """聚合表格数据：按维度列分组，对指标列求和
        
        Args:
            table_data: 包含headers和data的字典
            dim_columns: 维度列名列表
            ind_columns: 指标列名列表
            table_name: 表名（用于error标记）
            
        Returns:
            聚合后的表格数据（headers和data）
        """
        headers = table_data['headers']
        data = table_data['data']
        
        # 创建标准化列名到索引的映射
        normalized_header_map = {}
        for idx, header in enumerate(headers):
            normalized = self._normalize_string(header)
            normalized_header_map[normalized] = idx
        
        # 找到维度列和指标列的索引（使用标准化匹配）
        dim_indices = []
        for col in dim_columns:
            normalized_col = self._normalize_string(col)
            if normalized_col in normalized_header_map:
                dim_indices.append(normalized_header_map[normalized_col])
        
        ind_indices = []
        ind_names_in_table = []
        for col in ind_columns:
            normalized_col = self._normalize_string(col)
            if normalized_col in normalized_header_map:
                ind_indices.append(normalized_header_map[normalized_col])
                ind_names_in_table.append(col)
        
        # 按维度分组聚合
        groups = {}
        for row in data:
            # 构建维度键（使用标准化后的值，忽略下划线、空格和括号差异）
            dim_values = [row[i] if i < len(row) else None for i in dim_indices]
            dim_key = self._normalize_dimension_key(dim_values)
            
            if dim_key not in groups:
                groups[dim_key] = {}
                # 初始化所有指标为0
                for ind_name in ind_columns:
                    groups[dim_key][ind_name] = 0
            
            # 累加指标值
            for i, ind_idx in enumerate(ind_indices):
                ind_name = ind_names_in_table[i]
                try:
                    val = row[ind_idx] if ind_idx < len(row) else None
                    if val is not None and str(val).strip() != '':
                        groups[dim_key][ind_name] += float(val)
                except (ValueError, TypeError):
                    pass  # 忽略无法转换的值
        
        # 生成聚合后的表格
        agg_headers = dim_columns + ind_columns
        agg_data = []
        
        for dim_key, ind_values in groups.items():
            row = list(dim_key)  # 维度值
            # 添加指标值，如果不存在则显示error
            for ind_name in ind_columns:
                if ind_name in ind_names_in_table:
                    row.append(ind_values.get(ind_name, 0))
                else:
                    row.append('error')  # 该指标在原表中不存在
            agg_data.append(row)
        
        return {
            'headers': agg_headers,
            'data': agg_data,
            'actual_indicators': ind_names_in_table  # 原表中实际存在的指标
        }
    
    def _create_aggregate_result(self, output, agg_a, agg_b, 
                                 table_a_file, table_b_file,
                                 table_a_name, table_b_name,
                                 key_columns, diff_threshold):
        """创建聚合比对结果Excel（5个sheet）
        
        Args:
            output: 输出文件路径
            agg_a: 聚合后的表A数据
            agg_b: 聚合后的表B数据
            table_a_file: 表A源文件路径
            table_b_file: 表B源文件路径
            table_a_name: 表A名称
            table_b_name: 表B名称
            key_columns: 维度列数量
            diff_threshold: 差异阈值
        """
        wb = Workbook()
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        ERROR_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        HIGHLIGHT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        HEADER = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 找出A和B中不匹配的行（基于维度键）
        headers_a = agg_a['headers']
        headers_b = agg_b['headers']
        data_a = agg_a['data']
        data_b = agg_b['data']
        
        # 构建维度键索引
        a_keys = set()
        for row in data_a:
            key = tuple(row[:key_columns])
            a_keys.add(key)
        
        b_keys = set()
        for row in data_b:
            key = tuple(row[:key_columns])
            b_keys.add(key)
        
        # 找出不匹配的键
        only_in_a = a_keys - b_keys
        only_in_b = b_keys - a_keys
        
        # Sheet1: A表聚合（标红在A中存在但在B中不存在的行）
        ws1 = wb.create_sheet(f"{table_a_name}聚合")
        self._write_aggregated_sheet_with_highlight(
            ws1, agg_a, HEADER, ERROR_FILL, HIGHLIGHT_FILL, border, 
            key_columns, only_in_a  # 修复：应该标红只在A中的行
        )
        
        # Sheet2: B表聚合（标红在B中存在但在A中不存在的行）
        ws2 = wb.create_sheet(f"{table_b_name}聚合")
        self._write_aggregated_sheet_with_highlight(
            ws2, agg_b, HEADER, ERROR_FILL, HIGHLIGHT_FILL, border, 
            key_columns, only_in_b  # 修复：应该标红只在B中的行
        )
        
        # Sheet3: 比对结果（只比对公共指标）
        ws3 = wb.create_sheet("聚合比对结果")
        wb.active = ws3  # 设置为活动sheet
        
        # 找出公共指标（只考虑原表中实际存在的指标，忽略下划线、空格和括号差异）
        actual_indicators_a = agg_a.get('actual_indicators', [])
        actual_indicators_b = agg_b.get('actual_indicators', [])
        
        # 创建标准化指标名到原始名称的映射
        normalized_a = {self._normalize_string(ind): ind for ind in actual_indicators_a}
        normalized_b = {self._normalize_string(ind): ind for ind in actual_indicators_b}
        
        # 找出标准化后的公共指标
        common_normalized = set(normalized_a.keys()) & set(normalized_b.keys())
        
        # 使用A表中的原始指标名称（保持统一）
        common_indicators_list = sorted([normalized_a[norm] for norm in common_normalized])
        
        # 构建只包含公共指标的A和B表
        agg_a_common = self._filter_common_indicators(agg_a, key_columns, common_indicators_list)
        agg_b_common = self._filter_common_indicators(agg_b, key_columns, common_indicators_list)
        
        # 使用维度比对逻辑生成比对结果
        temp_output = output + '.temp.xlsx'
        self._create_dimension_result(
            temp_output, agg_a_common, agg_b_common, key_columns,
            table_a_name, table_b_name, diff_threshold,
            None, None  # 不复制源文件
        )
        
        # 从临时文件读取比对结果sheet
        temp_wb = load_workbook(temp_output)
        temp_ws = temp_wb.active
        
        # 复制比对结果到ws3
        for row in temp_ws.iter_rows():
            for cell in row:
                new_cell = ws3.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    try:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                    except:
                        pass
        
        # 复制列宽
        for col_letter in temp_ws.column_dimensions:
            if col_letter in temp_ws.column_dimensions:
                ws3.column_dimensions[col_letter].width = temp_ws.column_dimensions[col_letter].width
        
        # 冻结表头
        ws3.freeze_panes = 'A2'
        
        temp_wb.close()
        os.remove(temp_output)  # 删除临时文件
        
        # Sheet4和5: A和B源文件
        if table_a_file and os.path.exists(table_a_file):
            self._copy_sheet_from_file(wb, table_a_file, f"源文件_{table_a_name}")
        if table_b_file and os.path.exists(table_b_file):
            self._copy_sheet_from_file(wb, table_b_file, f"源文件_{table_b_name}")
        
        # 保存
        try:
            wb.save(output)
        except Exception as e:
            if sys.platform == 'win32':
                output_bytes = output.encode('utf-8')
                wb.save(output_bytes)
            else:
                raise e
        finally:
            wb.close()
    
    def _filter_common_indicators(self, table_data, key_columns, common_indicators):
        """过滤表格，只保留公共指标（使用标准化匹配）"""
        headers = table_data['headers']
        data = table_data['data']
        
        # 构建新表头：维度列 + 公共指标列
        new_headers = headers[:key_columns] + common_indicators
        
        # 创建标准化列名到索引的映射
        normalized_header_map = {}
        for idx, header in enumerate(headers):
            normalized = self._normalize_string(header)
            normalized_header_map[normalized] = idx
        
        # 找到公共指标在原表中的索引（使用标准化匹配）
        indicator_indices = []
        for ind in common_indicators:
            normalized_ind = self._normalize_string(ind)
            if normalized_ind in normalized_header_map:
                indicator_indices.append(normalized_header_map[normalized_ind])
        
        # 构建新数据
        new_data = []
        for row in data:
            new_row = list(row[:key_columns])  # 维度列
            for idx in indicator_indices:
                new_row.append(row[idx] if idx < len(row) else None)
            new_data.append(new_row)
        
        return {
            'headers': new_headers,
            'data': new_data
        }
    
    def _write_aggregated_sheet_with_highlight(self, ws, table_data, header_fill, 
                                               error_fill, highlight_fill, border, 
                                               key_columns, highlight_keys):
        """写入聚合后的sheet，并标红指定的行
        
        Args:
            ws: worksheet
            table_data: 表格数据
            header_fill: 表头填充色
            error_fill: error单元格填充色
            highlight_fill: 整行高亮填充色
            border: 边框
            key_columns: 维度列数量
            highlight_keys: 需要高亮的维度键集合
        """
        headers = table_data['headers']
        data = table_data['data']
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 冻结表头
        ws.freeze_panes = 'A2'
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            # 获取该行的维度键
            row_key = tuple(row_data[:key_columns])
            should_highlight = row_key in highlight_keys
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # 如果该行需要高亮
                if should_highlight:
                    cell.fill = highlight_fill
                    if isinstance(value, str) and 'error' in value.lower():
                        cell.font = Font(color="FF0000")
                # 否则只对error单元格标红
                elif isinstance(value, str) and 'error' in value.lower():
                    cell.fill = error_fill
                    cell.font = Font(color="FF0000")
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16
    
def main():
    print("=" * 50)
    print("Excel比对工具 - Web界面")
    print("=" * 50)
    print()
    
    if not OPENPYXL_OK:
        print("[警告] 缺少openpyxl库，请运行: pip install openpyxl")
        print()
    
    url = "http://localhost:{}".format(PORT)
    print("启动服务器: {}".format(url))
    print("按 Ctrl+C 停止服务器")
    print()
    
    # 自动打开浏览器
    threading.Timer(1, lambda: webbrowser.open(url)).start()
    
    # 启动服务器
    server = HTTPServer(('localhost', PORT), RequestHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务器已停止")
        server.shutdown()


if __name__ == '__main__':
    main()

