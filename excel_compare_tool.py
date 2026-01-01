#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel比对工具
Python 3.7.1 兼容

功能：
1. 从基准匹配列获取指标名称
2. 从两个横向数据源中匹配对应值
3. 计算差额和差异百分比
4. 根据差异百分比设置颜色（<1%绿色，>=1%红色）
"""

import argparse
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 颜色定义
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # 红色
HEADER_FILL = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid") # 灰色表头


def parse_number(value):
    """
    解析数字，处理各种格式（含逗号、空格等）
    返回 Decimal 或 None（如果无法解析）
    """
    if value is None:
        return None
    
    # 如果已经是数字类型
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    
    # 字符串处理
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ['error', '#value!', 'none', 'null', '']:
        return None
    
    try:
        # 移除逗号和空格
        cleaned = value_str.replace(',', '').replace(' ', '')
        return Decimal(cleaned)
    except Exception:
        return None


def read_base_column(file_path, sheet_name=None, column=1, start_row=2):
    """
    读取基准匹配列（纵向指标名称列表）
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，None则使用第一个
        column: 列号（1-based）
        start_row: 起始行号（跳过表头）
    
    Returns:
        list: 指标名称列表
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    names = []
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column).value
        if cell_value is not None and str(cell_value).strip():
            names.append(str(cell_value).strip())
    
    wb.close()
    return names


def read_horizontal_data(file_path, sheet_name=None, header_row=1, data_row=2):
    """
    读取横向数据（第一行为指标名称，第二行为对应值）
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        header_row: 表头行号
        data_row: 数据行号
    
    Returns:
        dict: {指标名称: 值}
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    data = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header is not None:
            header_str = str(header).strip()
            if header_str:
                value = ws.cell(row=data_row, column=col).value
                data[header_str] = value
    
    wb.close()
    return data


def calculate_difference(value_a, value_b):
    """
    计算差额：A - B
    
    Returns:
        Decimal 或 None
    """
    num_a = parse_number(value_a)
    num_b = parse_number(value_b)
    
    if num_a is None or num_b is None:
        return None
    
    return num_a - num_b


def calculate_percentage(difference, base_value):
    """
    计算差异百分比：差额 / 基准值 * 100
    保留4位小数
    
    Args:
        difference: 差额
        base_value: 基准值（用于计算百分比的分母，通常是B值）
    
    Returns:
        Decimal 或 None
    """
    if difference is None:
        return None
    
    base = parse_number(base_value)
    if base is None or base == 0:
        return None
    
    # 计算百分比，保留6位小数后再格式化
    percentage = (difference / base) * 100
    # 保留4位小数
    return percentage.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)


def format_decimal(value, decimal_places=4):
    """
    格式化Decimal为指定小数位
    """
    if value is None:
        return "error"
    
    format_str = '0.' + '0' * decimal_places
    return float(value.quantize(Decimal(format_str), rounding=ROUND_HALF_UP))


def create_output_excel(output_path, base_names, data_a, data_b):
    """
    创建输出Excel文件
    
    Args:
        output_path: 输出文件路径
        base_names: 基准指标名称列表
        data_a: 数据源A的字典
        data_b: 数据源B的字典
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "比对结果"
    
    # 设置表头
    headers = ["指标名称", "A", "B", "实际（差额）", "差异占比"]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 填充数据
    for row_idx, name in enumerate(base_names, 2):
        # 指标名称
        ws.cell(row=row_idx, column=1, value=name).border = thin_border
        
        # A值
        value_a = data_a.get(name)
        parsed_a = parse_number(value_a)
        if parsed_a is not None:
            ws.cell(row=row_idx, column=2, value=float(parsed_a)).border = thin_border
        else:
            cell_a = ws.cell(row=row_idx, column=2, value="error")
            cell_a.border = thin_border
        
        # B值
        value_b = data_b.get(name)
        parsed_b = parse_number(value_b)
        if parsed_b is not None:
            ws.cell(row=row_idx, column=3, value=float(parsed_b)).border = thin_border
        else:
            cell_b = ws.cell(row=row_idx, column=3, value="error")
            cell_b.border = thin_border
        
        # 差额
        difference = calculate_difference(value_a, value_b)
        if difference is not None:
            diff_value = format_decimal(difference, 4)
            ws.cell(row=row_idx, column=4, value=diff_value).border = thin_border
        else:
            cell_diff = ws.cell(row=row_idx, column=4, value="#VALUE!")
            cell_diff.border = thin_border
        
        # 差异占比（以B值为基准计算百分比）
        percentage = calculate_percentage(difference, value_b)
        cell_pct = ws.cell(row=row_idx, column=5)
        cell_pct.border = thin_border
        
        if percentage is not None:
            # 显示为百分比格式
            pct_str = "{}%".format(format_decimal(percentage, 6))
            cell_pct.value = pct_str
            
            # 根据百分比绝对值设置颜色
            abs_pct = abs(percentage)
            if abs_pct < 1:  # 绝对值 < 1%
                cell_pct.fill = GREEN_FILL
            elif abs_pct <= 100:  # 1% - 100%
                cell_pct.fill = RED_FILL
            # > 100% 不设置颜色（或可根据需求调整）
        else:
            cell_pct.value = "#VALUE!"
    
    # 调整列宽
    column_widths = [20, 20, 20, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 添加图例说明
    legend_row = len(base_names) + 4
    ws.cell(row=legend_row, column=5, value="差额绝对值大于1%")
    ws.cell(row=legend_row, column=6, value="大红色").fill = RED_FILL
    
    ws.cell(row=legend_row + 1, column=5, value="差额绝对值小于1%")
    ws.cell(row=legend_row + 1, column=6, value="").fill = GREEN_FILL
    
    # 保存文件
    wb.save(output_path)
    print("输出文件已保存至: {}".format(output_path))


def main():
    parser = argparse.ArgumentParser(
        description='Excel比对工具 - 比较两个数据源的差异',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例用法:
  python excel_compare_tool.py -b base.xlsx -a data_a.xlsx -c data_b.xlsx -o output.xlsx

文件格式说明:
  - 基准文件(-b): 纵向排列的指标名称列，第一列为指标名称
  - 数据源A(-a): 横向排列，第一行为指标名称，第二行为对应值
  - 数据源B(-c): 横向排列，第一行为指标名称，第二行为对应值
        '''
    )
    
    parser.add_argument('-b', '--base', required=True, help='基准匹配列Excel文件路径')
    parser.add_argument('-a', '--data-a', required=True, help='数据源A的Excel文件路径')
    parser.add_argument('-c', '--data-b', required=True, help='数据源B的Excel文件路径')
    parser.add_argument('-o', '--output', default='compare_output.xlsx', help='输出Excel文件路径')
    
    # 可选参数
    parser.add_argument('--base-sheet', default=None, help='基准文件的工作表名称')
    parser.add_argument('--base-column', type=int, default=1, help='基准文件中指标名称所在列（1-based）')
    parser.add_argument('--base-start-row', type=int, default=2, help='基准文件数据起始行（跳过表头）')
    
    parser.add_argument('--data-a-sheet', default=None, help='数据源A的工作表名称')
    parser.add_argument('--data-a-header-row', type=int, default=1, help='数据源A的表头行号')
    parser.add_argument('--data-a-data-row', type=int, default=2, help='数据源A的数据行号')
    
    parser.add_argument('--data-b-sheet', default=None, help='数据源B的工作表名称')
    parser.add_argument('--data-b-header-row', type=int, default=1, help='数据源B的表头行号')
    parser.add_argument('--data-b-data-row', type=int, default=2, help='数据源B的数据行号')
    
    args = parser.parse_args()
    
    print("正在读取基准匹配列: {}".format(args.base))
    base_names = read_base_column(
        args.base,
        sheet_name=args.base_sheet,
        column=args.base_column,
        start_row=args.base_start_row
    )
    print("  找到 {} 个指标".format(len(base_names)))
    
    print("正在读取数据源A: {}".format(args.data_a))
    data_a = read_horizontal_data(
        args.data_a,
        sheet_name=args.data_a_sheet,
        header_row=args.data_a_header_row,
        data_row=args.data_a_data_row
    )
    print("  读取到 {} 个数据项".format(len(data_a)))
    
    print("正在读取数据源B: {}".format(args.data_b))
    data_b = read_horizontal_data(
        args.data_b,
        sheet_name=args.data_b_sheet,
        header_row=args.data_b_header_row,
        data_row=args.data_b_data_row
    )
    print("  读取到 {} 个数据项".format(len(data_b)))
    
    print("正在生成比对结果...")
    create_output_excel(args.output, base_names, data_a, data_b)
    
    print("完成!")


if __name__ == '__main__':
    main()

